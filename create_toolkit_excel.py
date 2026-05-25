import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create new Real Estate Investment Toolkit Excel file
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create main sheets based on PDF content
sheets_config = [
    '00_Dashboard',
    '01_15_State_Framework',
    '02_County_Examples',
    '03_Resource_Finder',
    '04_Contact_Database',
    '05_Investment_Formulas',
    '06_Implementation',
    '07_Pro_Tips_Secrets',
    '08_Market_Intelligence',
    '09_Deal_Analysis',
    '10_Portfolio_Tracker'
]

for sheet_name in sheets_config:
    ws = wb.create_sheet(sheet_name)

    # Add header
    ws['A1'] = f'Real Estate Investment Master Toolkit - {sheet_name.replace("_", " ")}'
    header_font = Font(size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill

    # Merge cells for header
    ws.merge_cells('A1:Z1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add integration note
    ws['A3'] = 'INTEGRATED WITH: Land_and_Build_Flipping_System_v2.xlsx'
    ws['A3'].font = Font(size=10, italic=True)
    ws['A3'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    ws.merge_cells('A3:Z3')

# Populate Dashboard with key metrics and integration points
dashboard = wb['00_Dashboard']
dashboard['A5'] = 'Key Performance Indicators'
dashboard['A5'].font = Font(size=14, bold=True)

dashboard['A7'] = 'Total Deals Analyzed:'
dashboard['B7'] = '=COUNTIF(\'09_Deal_Analysis\'!A:A, \"<>\")-1'

dashboard['A8'] = 'Active Markets:'
dashboard['B8'] = '=COUNTA(\'01_15_State_Framework\'!A:A)-1'

dashboard['A9'] = 'Contact Database Size:'
dashboard['B9'] = '=COUNTA(\'04_Contact_Database\'!A:A)-1'

dashboard['A10'] = 'Portfolio Value:'
dashboard['B10'] = '=SUM(\'10_Portfolio_Tracker\'!D:D)'

# Add state framework data
state_ws = wb['01_15_State_Framework']
state_ws['A5'] = 'State'
state_ws['B5'] = 'Market Status'
state_ws['C5'] = 'Key Counties'
state_ws['D5'] = 'Investment Focus'
state_ws['E5'] = 'Resources Available'

# Sample state data
states_data = [
    ['Texas', 'Hot', 'Harris, Dallas, Tarrant', 'Multi-family, Land', 'Yes'],
    ['Florida', 'Hot', 'Miami-Dade, Broward', 'Condo, Single-family', 'Yes'],
    ['Arizona', 'Growing', 'Maricopa, Pima', 'Single-family, Land', 'Yes'],
    ['North Carolina', 'Emerging', 'Mecklenburg, Wake', 'Single-family', 'Yes'],
    ['Georgia', 'Hot', 'Fulton, Gwinnett', 'Multi-family', 'Yes']
]

for i, row in enumerate(states_data, 6):
    for j, value in enumerate(row):
        state_ws.cell(row=i, column=j+1, value=value)

# Add investment formulas
formulas_ws = wb['05_Investment_Formulas']
formulas_ws['A5'] = 'Formula Name'
formulas_ws['B5'] = 'Formula'
formulas_ws['C5'] = 'Description'
formulas_ws['D5'] = 'Example'

formulas_data = [
    ['Cap Rate', '=(Annual NOI / Property Value) * 100', 'Capitalization Rate', '=($50,000 / $500,000) * 100 = 10%'],
    ['Cash-on-Cash Return', '=(Annual Cash Flow / Initial Investment) * 100', 'Cash flow return on investment', '=($25,000 / $100,000) * 100 = 25%'],
    ['IRR', '=IRR(Cash Flow Array)', 'Internal Rate of Return', 'Complex calculation requiring cash flows'],
    ['DSCR', '=Annual NOI / Annual Debt Service', 'Debt Service Coverage Ratio', '=($60,000 / $45,000) = 1.33'],
    ['NOI', '=Gross Income - Operating Expenses', 'Net Operating Income', '=($100,000 - $40,000) = $60,000']
]

for i, row in enumerate(formulas_data, 6):
    for j, value in enumerate(row):
        formulas_ws.cell(row=i, column=j+1, value=value)

# Add pro tips section
tips_ws = wb['07_Pro_Tips_Secrets']
tips_ws['A5'] = 'Category'
tips_ws['B5'] = 'Tip/Secret'
tips_ws['C5'] = 'Impact Level'

tips_data = [
    ['Market Analysis', 'Always verify comps within 1-mile radius', 'High'],
    ['Due Diligence', 'Never skip environmental assessment', 'Critical'],
    ['Negotiation', 'Always have 3 exit strategies', 'High'],
    ['Financing', 'Shop rates 30 days before closing', 'Medium'],
    ['Property Management', 'Budget 1% of property value for maintenance', 'High']
]

for i, row in enumerate(tips_data, 6):
    for j, value in enumerate(row):
        tips_ws.cell(row=i, column=j+1, value=value)

# Save the new toolkit
wb.save('Real_Estate_Investment_Toolkit.xlsx')
print('Created Real_Estate_Investment_Toolkit.xlsx with integrated structure and sample data')