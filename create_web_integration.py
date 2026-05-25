import openpyxl
import json
import os
from datetime import datetime

def create_web_integration():
    """Create web interface integration with the Excel toolkits"""

    # Load the integrated toolkits
    try:
        toolkit_wb = openpyxl.load_workbook('Real_Estate_Investment_Toolkit.xlsx')
        existing_wb = openpyxl.load_workbook('Land_and_Build_Flipping_System_v2.xlsx')
    except FileNotFoundError as e:
        print(f"Error loading workbooks: {e}")
        return

    print("Loading integrated Excel toolkits for web interface...")

    # Extract key data for web integration
    integration_data = {
        'timestamp': datetime.now().isoformat(),
        'states': [],
        'formulas': [],
        'tips': [],
        'market_data': {},
        'contact_count': 0
    }

    # Extract state framework data
    if '01_15_State_Framework' in toolkit_wb.sheetnames:
        state_ws = toolkit_wb['01_15_State_Framework']
        for row in range(6, state_ws.max_row + 1):
            state = state_ws.cell(row=row, column=1).value
            status = state_ws.cell(row=row, column=2).value
            counties = state_ws.cell(row=row, column=3).value
            focus = state_ws.cell(row=row, column=4).value

            if state:
                integration_data['states'].append({
                    'name': state,
                    'status': status or 'Unknown',
                    'counties': counties or '',
                    'focus': focus or '',
                    'resources_available': True
                })

    # Extract investment formulas
    if '05_Investment_Formulas' in toolkit_wb.sheetnames:
        formulas_ws = toolkit_wb['05_Investment_Formulas']
        for row in range(6, formulas_ws.max_row + 1):
            name = formulas_ws.cell(row=row, column=1).value
            formula = formulas_ws.cell(row=row, column=2).value
            description = formulas_ws.cell(row=row, column=3).value

            if name and formula:
                integration_data['formulas'].append({
                    'name': name,
                    'formula': formula,
                    'description': description or ''
                })

    # Extract pro tips
    if '07_Pro_Tips_Secrets' in toolkit_wb.sheetnames:
        tips_ws = toolkit_wb['07_Pro_Tips_Secrets']
        for row in range(6, tips_ws.max_row + 1):
            category = tips_ws.cell(row=row, column=1).value
            tip = tips_ws.cell(row=row, column=2).value
            impact = tips_ws.cell(row=row, column=3).value

            if tip:
                integration_data['tips'].append({
                    'category': category or 'General',
                    'tip': tip,
                    'impact': impact or 'Medium'
                })

    # Count contacts from existing CRM
    if 'CRM' in existing_wb.sheetnames:
        crm_ws = existing_wb['CRM']
        integration_data['contact_count'] = max(0, crm_ws.max_row - 1)  # Subtract header row

    # Create market intelligence summary
    integration_data['market_data'] = {
        'total_states': len(integration_data['states']),
        'hot_markets': len([s for s in integration_data['states'] if s['status'] == 'Hot']),
        'growing_markets': len([s for s in integration_data['states'] if s['status'] == 'Growing']),
        'total_formulas': len(integration_data['formulas']),
        'pro_tips_count': len(integration_data['tips'])
    }

    # Save integration data for web interface
    with open('web_integration_data.json', 'w') as f:
        json.dump(integration_data, f, indent=2)

    print(f"Created web_integration_data.json with {len(integration_data['states'])} states, {len(integration_data['formulas'])} formulas, {len(integration_data['tips'])} tips")

    # Create enhanced web server integration
    web_integration_code = '''
# Web Integration for Real Estate Investment Toolkit
# Add this to your web_server.py to integrate toolkit data

def load_toolkit_data():
    """Load integrated toolkit data for web interface"""
    try:
        with open('web_integration_data.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            'states': [],
            'formulas': [],
            'tips': [],
            'market_data': {},
            'contact_count': 0
        }

# Enhanced route for toolkit dashboard
@app.route('/toolkit-dashboard')
def toolkit_dashboard():
    toolkit_data = load_toolkit_data()
    return render_template('toolkit_dashboard.html',
                         states=toolkit_data['states'],
                         formulas=toolkit_data['formulas'],
                         tips=toolkit_data['tips'],
                         market_data=toolkit_data['market_data'],
                         contact_count=toolkit_data['contact_count'])

# Enhanced deal finding with toolkit integration
@app.route('/enhanced-deals')
def enhanced_deals():
    # Load toolkit data
    toolkit_data = load_toolkit_data()

    # Get AI-powered deals
    scraper = RealEstateLeadScraper()
    deals = scraper.auto_find_deals()

    # Enhance deals with toolkit data
    enhanced_deals = []
    for deal in deals:
        # Add state-specific insights
        deal_state = deal.get('state', '')
        state_info = next((s for s in toolkit_data['states'] if s['name'] == deal_state), None)

        if state_info:
            deal['market_status'] = state_info['status']
            deal['investment_focus'] = state_info['focus']

        # Add formula calculations
        if 'price' in deal and 'noi' in deal:
            cap_rate = (deal['noi'] / deal['price']) * 100 if deal['price'] > 0 else 0
            deal['cap_rate'] = round(cap_rate, 2)

        enhanced_deals.append(deal)

    return render_template('enhanced_deals.html', deals=enhanced_deals)
'''

    with open('web_integration_enhancement.py', 'w') as f:
        f.write(web_integration_code)

    print("Created web_integration_enhancement.py for Flask integration")

    # Create HTML template for toolkit dashboard
    html_template = '''
<!DOCTYPE html>
<html>
<head>
    <title>Real Estate Investment Toolkit Dashboard</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .header { background: #4F81BD; color: white; padding: 20px; text-align: center; }
        .metric-card { background: #f0f0f0; padding: 15px; margin: 10px; border-radius: 5px; display: inline-block; min-width: 200px; }
        .state-card { background: #e6f3ff; padding: 10px; margin: 5px; border-radius: 3px; }
        .hot-market { border-left: 5px solid #ff6600; }
        .growing-market { border-left: 5px solid #66cc00; }
        .formula-item { background: #fff2e6; padding: 10px; margin: 5px 0; border-radius: 3px; }
        .tip-item { background: #e6ffe6; padding: 10px; margin: 5px 0; border-radius: 3px; }
    </style>
</head>
<body>
    <div class="header">
        <h1>Real Estate Investment Master Toolkit</h1>
        <p>Integrated Professional Analysis System</p>
    </div>

    <h2>Market Intelligence Dashboard</h2>
    <div class="metric-card">
        <h3>Total States Covered: {{ market_data.total_states }}</h3>
    </div>
    <div class="metric-card">
        <h3>Hot Markets: {{ market_data.hot_markets }}</h3>
    </div>
    <div class="metric-card">
        <h3>Growing Markets: {{ market_data.growing_markets }}</h3>
    </div>
    <div class="metric-card">
        <h3>Contact Database: {{ contact_count }}</h3>
    </div>

    <h2>State Framework</h2>
    {% for state in states %}
    <div class="state-card {{ 'hot-market' if state.status == 'Hot' else 'growing-market' if state.status == 'Growing' else '' }}">
        <h4>{{ state.name }} - {{ state.status }}</h4>
        <p><strong>Key Counties:</strong> {{ state.counties }}</p>
        <p><strong>Investment Focus:</strong> {{ state.investment_focus }}</p>
    </div>
    {% endfor %}

    <h2>Investment Formulas</h2>
    {% for formula in formulas %}
    <div class="formula-item">
        <h4>{{ formula.name }}</h4>
        <p><strong>Formula:</strong> {{ formula.formula }}</p>
        <p><strong>Description:</strong> {{ formula.description }}</p>
    </div>
    {% endfor %}

    <h2>Pro Tips & Secrets</h2>
    {% for tip in tips %}
    <div class="tip-item">
        <h4>{{ tip.category }} - {{ tip.impact }} Impact</h4>
        <p>{{ tip.tip }}</p>
    </div>
    {% endfor %}

    <br><br>
    <a href="/">Back to Main Dashboard</a>
</body>
</html>
'''

    # Save HTML template
    os.makedirs('templates', exist_ok=True)
    with open('templates/toolkit_dashboard.html', 'w') as f:
        f.write(html_template)

    print("Created templates/toolkit_dashboard.html")

    print("\n" + "="*60)
    print("WEB INTEGRATION COMPLETE!")
    print("="*60)
    print("Files created:")
    print("• web_integration_data.json (toolkit data for web)")
    print("• web_integration_enhancement.py (Flask integration code)")
    print("• templates/toolkit_dashboard.html (web dashboard template)")
    print("\nTo integrate with web server:")
    print("1. Copy the route functions from web_integration_enhancement.py to web_server.py")
    print("2. Add json import to web_server.py")
    print("3. The /toolkit-dashboard route will show integrated toolkit data")
    print("4. Enhanced deal finding available at /enhanced-deals")

if __name__ == '__main__':
    create_web_integration()