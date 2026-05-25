import openpyxl

try:
    wb = openpyxl.load_workbook('Land_and_Build_Flipping_System_v2.xlsx')
    print('Existing Excel file sheets:')
    for sheet in wb.sheetnames:
        print(f'  - {sheet}')
    print(f'Total sheets: {len(wb.sheetnames)}')

    # Check the main CRM sheet
    if '09_Leads_CRM' in wb.sheetnames:
        ws = wb['09_Leads_CRM']
        print(f'\n09_Leads_CRM sheet has {ws.max_row} rows and {ws.max_column} columns')
        # Check first few rows
        for row in range(1, min(6, ws.max_row + 1)):
            values = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                values.append(str(cell_value) if cell_value is not None else '')
            print(f'Row {row}: {values}')

except Exception as e:
    print(f'Error reading Excel file: {e}')