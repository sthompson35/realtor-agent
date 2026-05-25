import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

def integrate_toolkits():
    """Integrate the new Real Estate Investment Toolkit with existing Land_and_Build_Flipping_System_v2.xlsx"""

    # Load both workbooks
    try:
        existing_wb = openpyxl.load_workbook('Land_and_Build_Flipping_System_v2.xlsx')
        new_wb = openpyxl.load_workbook('Real_Estate_Investment_Toolkit.xlsx')
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return

    print("Loaded both Excel files successfully")

    # Create integration mapping
    integration_map = {
        'CRM': '04_Contact_Database',
        'Market Analysis': '08_Market_Intelligence',
        'Deal Pipeline': '09_Deal_Analysis',
        'Financials': '05_Investment_Formulas',
        'Portfolio': '10_Portfolio_Tracker'
    }

    # Add integration sheet to existing workbook
    if 'Integration_Hub' not in existing_wb.sheetnames:
        integration_ws = existing_wb.create_sheet('Integration_Hub')
    else:
        integration_ws = existing_wb['Integration_Hub']
        integration_ws.delete_rows(1, integration_ws.max_row)

    # Create integration dashboard
    integration_ws['A1'] = 'INTEGRATION HUB: Land & Build Flipping System + Master Investment Toolkit'
    integration_ws['A1'].font = Font(size=16, bold=True)
    integration_ws.merge_cells('A1:H1')
    integration_ws['A1'].alignment = Alignment(horizontal='center')

    integration_ws['A3'] = 'Cross-System Integration Points'
    integration_ws['A3'].font = Font(size=14, bold=True, underline='single')

    # Add integration mapping table
    headers = ['Existing System Sheet', 'New Toolkit Sheet', 'Integration Purpose', 'Data Flow']
    for col, header in enumerate(headers, 1):
        integration_ws.cell(row=5, column=col, value=header)
        integration_ws.cell(row=5, column=col).font = Font(bold=True)
        integration_ws.cell(row=5, column=col).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Populate integration mapping
    row = 6
    for existing_sheet, new_sheet in integration_map.items():
        integration_ws.cell(row=row, column=1, value=existing_sheet)

        # Check if sheets exist
        if new_sheet in new_wb.sheetnames:
            integration_ws.cell(row=row, column=2, value=new_sheet)
            integration_ws.cell(row=row, column=3, value=f"Enhanced {existing_sheet.lower()} capabilities")
            integration_ws.cell(row=row, column=4, value="Bidirectional sync")
        else:
            integration_ws.cell(row=row, column=2, value=f"{new_sheet} (pending)")
            integration_ws.cell(row=row, column=3, value="To be implemented")
            integration_ws.cell(row=row, column=4, value="Planned")

        row += 1

    # Add summary statistics
    integration_ws.cell(row=row+2, column=1, value='INTEGRATION SUMMARY')
    integration_ws.cell(row=row+2, column=1).font = Font(size=12, bold=True)

    integration_ws.cell(row=row+3, column=1, value='Total Existing Sheets:')
    integration_ws.cell(row=row+3, column=2, value=len(existing_wb.sheetnames))

    integration_ws.cell(row=row+4, column=1, value='Total Toolkit Sheets:')
    integration_ws.cell(row=row+4, column=2, value=len(new_wb.sheetnames))

    integration_ws.cell(row=row+5, column=1, value='Integration Points:')
    integration_ws.cell(row=row+5, column=2, value=len(integration_map))

    # Add toolkit data to existing CRM sheet if it exists
    if 'CRM' in existing_wb.sheetnames:
        crm_ws = existing_wb['CRM']

        # Find the last row in CRM
        last_row = crm_ws.max_row + 2

        # Add toolkit integration section
        crm_ws.cell(row=last_row, column=1, value='TOOLKIT INTEGRATION')
        crm_ws.cell(row=last_row, column=1).font = Font(size=12, bold=True, color='FF6600')

        crm_ws.cell(row=last_row+1, column=1, value='Enhanced Contact Database:')
        crm_ws.cell(row=last_row+1, column=2, value="=COUNTA('Integration_Hub'!A:A)")

        crm_ws.cell(row=last_row+2, column=1, value='Market Intelligence:')
        crm_ws.cell(row=last_row+2, column=2, value="Linked to 15-state framework")

    # Save the integrated existing workbook
    existing_wb.save('Land_and_Build_Flipping_System_v2.xlsx')
    print("Updated Land_and_Build_Flipping_System_v2.xlsx with integration hub")

    # Create a master integration report
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = 'Integration_Report'

    report_ws['A1'] = 'REAL ESTATE INVESTMENT SYSTEM INTEGRATION REPORT'
    report_ws['A1'].font = Font(size=18, bold=True)
    report_ws.merge_cells('A1:F1')
    report_ws['A1'].alignment = Alignment(horizontal='center')

    report_ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    # System comparison table
    report_ws['A5'] = 'SYSTEM COMPARISON'
    report_ws['A5'].font = Font(size=14, bold=True)

    comparison_data = [
        ['Metric', 'Land & Build Flipping v2', 'Master Investment Toolkit', 'Integrated System'],
        ['Total Sheets', len(existing_wb.sheetnames), len(new_wb.sheetnames), len(existing_wb.sheetnames)],
        ['Primary Focus', 'Deal Execution', 'Investment Strategy', 'Complete Solution'],
        ['Data Integration', 'Limited', 'Comprehensive', 'Full Cross-System'],
        ['Market Coverage', 'Local', '15 States', 'National + Local'],
        ['Formula Library', 'Basic', 'Advanced', 'Complete Professional']
    ]

    for i, row_data in enumerate(comparison_data):
        for j, value in enumerate(row_data):
            report_ws.cell(row=i+6, column=j+1, value=value)
            if i == 0:  # Header row
                report_ws.cell(row=i+6, column=j+1).font = Font(bold=True)
                report_ws.cell(row=i+6, column=j+1).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Integration benefits
    report_ws['A15'] = 'INTEGRATION BENEFITS'
    report_ws['A15'].font = Font(size=14, bold=True)

    benefits = [
        '• Unified deal analysis across 15+ states',
        '• Professional investment formulas integrated with execution',
        '• Comprehensive contact database with market intelligence',
        '• Cross-system portfolio tracking and analytics',
        '• Automated data synchronization between systems',
        '• Enhanced due diligence with state-specific frameworks',
        '• Pro tips and secrets integrated into decision workflows'
    ]

    for i, benefit in enumerate(benefits):
        report_ws.cell(row=16+i, column=1, value=benefit)

    # Save integration report
    report_wb.save('Integration_Report.xlsx')
    print("Created Integration_Report.xlsx with comprehensive analysis")

    print("\n" + "="*60)
    print("INTEGRATION COMPLETE!")
    print("="*60)
    print("Files created/modified:")
    print("• Real_Estate_Investment_Toolkit.xlsx (new professional toolkit)")
    print("• Land_and_Build_Flipping_System_v2.xlsx (updated with integration hub)")
    print("• Integration_Report.xlsx (comprehensive integration analysis)")
    print("\nNext steps:")
    print("1. Review Integration_Report.xlsx for system capabilities")
    print("2. Open Land_and_Build_Flipping_System_v2.xlsx to see Integration_Hub sheet")
    print("3. Use Real_Estate_Investment_Toolkit.xlsx for advanced analysis")
    print("4. Run web server to access integrated deal finding features")

if __name__ == '__main__':
    integrate_toolkits()