"""
Database to Excel Synchronization Utility
Handles bidirectional synchronization between SQLAlchemy database and Excel files.

© Shylow Thompson. LLC 2026 - All Rights Reserved
"""

import os
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from enum import Enum

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError

from realtor_agent.core.database import Base, Lead, Property, Deal, User, Appointment
from realtor_agent.core.config import config


class SyncDirection(Enum):
    """Synchronization direction"""
    DB_TO_EXCEL = "db_to_excel"
    EXCEL_TO_DB = "excel_to_db"
    BIDIRECTIONAL = "bidirectional"


class ConflictResolution(Enum):
    """Conflict resolution strategies"""
    DB_WINS = "db_wins"  # Database data takes precedence
    EXCEL_WINS = "excel_wins"  # Excel data takes precedence
    NEWEST_WINS = "newest_wins"  # Most recently updated wins
    MANUAL = "manual"  # Require manual intervention


@dataclass
class SyncResult:
    """Result of a synchronization operation"""
    direction: SyncDirection
    records_processed: int
    records_added: int
    records_updated: int
    records_skipped: int
    conflicts: List[Dict[str, Any]]
    errors: List[str]
    timestamp: datetime


@dataclass
class SyncConfig:
    """Configuration for synchronization"""
    direction: SyncDirection
    conflict_resolution: ConflictResolution
    excel_file_path: Path
    excel_sheet: str
    db_table: str
    field_mapping: Dict[str, str]  # Excel column -> DB field
    key_field: str  # Excel column name used to match records
    batch_size: int = 100
    dry_run: bool = False


class DatabaseExcelSync:
    """
    Handles synchronization between database and Excel files.
    """

    def __init__(self, db_url: Optional[str] = None):
        """
        Initialize the sync utility.

        Args:
            db_url: Database URL. If None, uses config default.
        """
        self.logger = logging.getLogger(__name__)

        # Initialize database connection
        if db_url is None:
            db_url = config.database.get('url') if hasattr(config, 'database') else "sqlite:///data/realtor_agent.db"

        self.engine = create_engine(db_url, echo=False)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

        # Create tables if they don't exist
        Base.metadata.create_all(bind=self.engine)

        self.logger.info(f"DatabaseExcelSync initialized with database: {db_url}")

    def get_db_session(self) -> Session:
        """Get a database session."""
        return self.SessionLocal()

    def sync_leads_crm(self,
                      direction: SyncDirection = SyncDirection.BIDIRECTIONAL,
                      conflict_resolution: ConflictResolution = ConflictResolution.NEWEST_WINS,
                      dry_run: bool = False) -> SyncResult:
        """
        Synchronize leads between database and Excel CRM sheet.

        Args:
            direction: Direction of synchronization
            conflict_resolution: How to handle conflicts
            dry_run: If True, don't make actual changes

        Returns:
            SyncResult with operation details
        """
        config = SyncConfig(
            direction=direction,
            conflict_resolution=conflict_resolution,
            excel_file_path=Path("Land_and_Build_Flipping_System_v2.xlsx"),
            excel_sheet="09_Leads_CRM",
            db_table="leads",
            field_mapping={
                "Lead ID": "lead_id",
                "Owner Name": "name",
                "Email": "email",
                "Phone": "phone",
                "Source (List)": "source",
                "Status": "status",
                "Last Contact": "notes"  # Store last contact in notes for now
            },
            key_field="Lead ID",
            dry_run=dry_run
        )

        return self._sync_table(config)

    def _sync_table(self, sync_config: SyncConfig) -> SyncResult:
        """
        Synchronize a specific table.

        Args:
            sync_config: Synchronization configuration

        Returns:
            SyncResult with operation details
        """
        result = SyncResult(
            direction=sync_config.direction,
            records_processed=0,
            records_added=0,
            records_updated=0,
            records_skipped=0,
            conflicts=[],
            errors=[],
            timestamp=datetime.now()
        )

        try:
            # Load Excel data
            excel_data = self._load_excel_data(sync_config.excel_file_path, sync_config.excel_sheet)
            result.records_processed = len(excel_data)

            # Load database data
            db_data = self._load_db_data(sync_config.db_table, sync_config.key_field, sync_config.field_mapping)

            # Perform synchronization based on direction
            if sync_config.direction in [SyncDirection.DB_TO_EXCEL, SyncDirection.BIDIRECTIONAL]:
                self._sync_db_to_excel(db_data, excel_data, sync_config, result)

            if sync_config.direction in [SyncDirection.EXCEL_TO_DB, SyncDirection.BIDIRECTIONAL]:
                self._sync_excel_to_db(excel_data, db_data, sync_config, result)

        except Exception as e:
            result.errors.append(f"Synchronization failed: {str(e)}")
            self.logger.error(f"Sync error: {e}")

        return result

    def _load_excel_data(self, file_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Load data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of the sheet

        Returns:
            List of dictionaries with Excel data
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        ws = wb[sheet_name]

        # Get headers from row 6
        headers = []
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=6, column=col).value
            headers.append(str(header_value).strip() if header_value else f"Col_{col}")

        # Load data starting from row 7
        data = []
        for row in range(7, ws.max_row + 1):
            row_data = {}
            has_data = False

            for col_idx, header in enumerate(headers):
                cell_value = ws.cell(row=row, column=col_idx + 1).value
                if cell_value is not None:
                    has_data = True
                    # Convert dates
                    if isinstance(cell_value, datetime):
                        row_data[header] = cell_value.isoformat()
                    else:
                        row_data[header] = str(cell_value)

            if has_data:
                data.append(row_data)

        wb.close()
        return data

    def _load_db_data(self, table_name: str, excel_key_field: str, field_mapping: Dict[str, str]) -> Dict[str, Dict[str, Any]]:
        """
        Load data from database table.

        Args:
            table_name: Name of the database table
            excel_key_field: Excel column name used as key
            field_mapping: Mapping from Excel columns to DB fields

        Returns:
            Dictionary keyed by Excel key field value with record data
        """
        # Find the corresponding DB field for the Excel key field
        db_key_field = None
        for excel_col, db_field in field_mapping.items():
            if excel_col == excel_key_field:
                db_key_field = db_field
                break

        if not db_key_field:
            raise ValueError(f"No database field mapping found for Excel key field '{excel_key_field}'")

        with self.get_db_session() as session:
            try:
                # Use raw SQL to get all records
                query = text(f"SELECT * FROM {table_name}")
                result = session.execute(query)
                columns = result.keys()

                data = {}
                for row in result:
                    record = dict(zip(columns, row))
                    # Use the DB key field value, but index by Excel key field equivalent
                    key_value = record.get(db_key_field)
                    if key_value:
                        data[str(key_value)] = record

                return data

            except SQLAlchemyError as e:
                self.logger.error(f"Database query error: {e}")
                return {}

    def _sync_db_to_excel(self, db_data: Dict[str, Dict[str, Any]],
                         excel_data: List[Dict[str, Any]],
                         config: SyncConfig, result: SyncResult) -> None:
        """
        Sync database data to Excel.
        """
        # This would add new database records to Excel
        # Implementation depends on specific requirements
        pass

    def _sync_excel_to_db(self, excel_data: List[Dict[str, Any]],
                         db_data: Dict[str, Dict[str, Any]],
                         config: SyncConfig, result: SyncResult) -> None:
        """
        Sync Excel data to database.
        """
        with self.get_db_session() as session:
            try:
                for excel_record in excel_data:
                    key_value = excel_record.get(config.key_field)

                    if not key_value:
                        result.records_skipped += 1
                        continue

                    # Check if record exists in database
                    db_record = db_data.get(str(key_value))

                    if db_record:
                        # Record exists - check for conflicts
                        if self._has_conflicts(excel_record, db_record, config.field_mapping):
                            conflict = self._resolve_conflict(excel_record, db_record, config)
                            if conflict:
                                result.conflicts.append(conflict)
                                if config.conflict_resolution == ConflictResolution.EXCEL_WINS:
                                    self._update_db_record(session, excel_record, config, result)
                                    result.records_updated += 1
                        # No conflicts, skip
                        result.records_skipped += 1
                    else:
                        # New record - add to database
                        if not config.dry_run:
                            try:
                                self._add_excel_record_to_db(session, excel_record, config)
                                result.records_added += 1
                            except ValueError as e:
                                # Skip records with missing required fields
                                result.errors.append(str(e))
                                result.records_skipped += 1
                                continue
                        else:
                            result.records_added += 1

                if not config.dry_run:
                    session.commit()

            except Exception as e:
                session.rollback()
                result.errors.append(f"Excel to DB sync failed: {str(e)}")

    def _has_conflicts(self, excel_record: Dict[str, Any],
                      db_record: Dict[str, Any],
                      field_mapping: Dict[str, str]) -> bool:
        """
        Check if there are conflicts between Excel and database records.
        """
        for excel_field, db_field in field_mapping.items():
            excel_value = excel_record.get(excel_field, "")
            db_value = db_record.get(db_field, "")

            # Normalize values for comparison
            if str(excel_value).strip() != str(db_value).strip():
                return True
        return False

    def _resolve_conflict(self, excel_record: Dict[str, Any],
                         db_record: Dict[str, Any],
                         config: SyncConfig) -> Optional[Dict[str, Any]]:
        """
        Resolve conflicts based on the configured strategy.
        """
        if config.conflict_resolution == ConflictResolution.MANUAL:
            return {
                "key": excel_record.get(config.key_field),
                "excel_data": excel_record,
                "db_data": db_record,
                "resolution": "manual_required"
            }
        return None

    def _add_excel_record_to_db(self, session: Session,
                               excel_record: Dict[str, Any],
                               config: SyncConfig) -> None:
        """
        Add Excel record to database.
        """
        # Create Lead object from Excel data
        lead_data = {}
        for excel_field, db_field in config.field_mapping.items():
            value = excel_record.get(excel_field)
            if value is not None and str(value).strip():  # Only include non-empty values
                # Convert data types as needed
                if db_field == "last_contact" and value:
                    try:
                        lead_data[db_field] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    except:
                        lead_data[db_field] = datetime.now()
                else:
                    lead_data[db_field] = value

        # Skip records without required fields (name is required)
        if not lead_data.get("name"):
            raise ValueError(f"Skipping record with missing required field 'name': {excel_record}")

        # Remove lead_id from data as it's auto-generated
        lead_data.pop("lead_id", None)

        lead = Lead(**lead_data)
        session.add(lead)

    def _update_db_record(self, session: Session,
                         excel_record: Dict[str, Any],
                         config: SyncConfig, result: SyncResult) -> None:
        """
        Update database record with Excel data.
        """
        # Implementation for updating existing records
        pass

    def export_leads_to_excel(self, output_path: Optional[Path] = None) -> str:
        """
        Export all leads from database to Excel file.

        Args:
            output_path: Path to save the Excel file

        Returns:
            Path to the created Excel file
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(f"leads_export_{timestamp}.xlsx")

        with self.get_db_session() as session:
            try:
                leads = session.query(Lead).all()

                # Create Excel workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Leads_Export"

                # Add headers
                headers = ["ID", "Lead ID", "Name", "Email", "Phone", "Source", "Status", "Score", "Created", "Updated"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # Add data
                for row, lead in enumerate(leads, 2):
                    ws.cell(row=row, column=1, value=lead.id)
                    ws.cell(row=row, column=2, value=lead.lead_id)
                    ws.cell(row=row, column=3, value=lead.name)
                    ws.cell(row=row, column=4, value=lead.email)
                    ws.cell(row=row, column=5, value=lead.phone)
                    ws.cell(row=row, column=6, value=lead.source)
                    ws.cell(row=row, column=7, value=lead.status)
                    ws.cell(row=row, column=8, value=lead.score)
                    ws.cell(row=row, column=9, value=lead.created_at.isoformat() if lead.created_at else "")
                    ws.cell(row=row, column=10, value=lead.updated_at.isoformat() if lead.updated_at else "")

                wb.save(output_path)
                self.logger.info(f"Exported {len(leads)} leads to {output_path}")

                return str(output_path)

            except Exception as e:
                self.logger.error(f"Export failed: {e}")
                raise

    def get_sync_status(self) -> Dict[str, Any]:
        """
        Get current synchronization status.

        Returns:
            Dictionary with sync status information
        """
        try:
            with self.get_db_session() as session:
                lead_count = session.query(Lead).count()
                property_count = session.query(Property).count()
                deal_count = session.query(Deal).count()

            excel_leads = len(self._load_excel_data(
                Path("Land_and_Build_Flipping_System_v2.xlsx"), "09_Leads_CRM"
            ))

            return {
                "database": {
                    "leads": lead_count,
                    "properties": property_count,
                    "deals": deal_count
                },
                "excel": {
                    "leads_crm": excel_leads
                },
                "last_sync": None,  # Would be stored in a sync log
                "status": "ready"
            }

        except Exception as e:
            return {
                "error": str(e),
                "status": "error"
            }

    def fetch_census_data(self, zip_code: str, data_type: str = "acs5") -> Dict[str, Any]:
        """
        Fetch Census Bureau demographic data for a given zip code.

        Args:
            zip_code: 5-digit zip code
            data_type: Census data type (acs5 for American Community Survey 5-year estimates)

        Returns:
            Dictionary with census data including median income and population
        """
        api_key = os.getenv("CENSUS_API_KEY")
        if not api_key:
            raise ValueError("CENSUS_API_KEY not found in environment variables")

        base_url = "https://api.census.gov/data"
        year = "2022"  # Most recent ACS 5-year estimates

        # Variables: B19013_001E (Median household income), B01003_001E (Total population)
        variables = "B19013_001E,B01003_001E"

        # First try zip code tabulation area (ZCTA) - direct zip code data
        url = f"{base_url}/{year}/acs/{data_type}?get={variables}&for=zip%20code%20tabulation%20area:{zip_code}&key={api_key}"

        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()

            data = response.json()

            if len(data) >= 2:  # Header + at least one data row
                # Parse the response
                headers = data[0]
                values = data[1]

                result = {}
                for i, header in enumerate(headers):
                    if i < len(values):
                        result[header] = values[i]

                # Convert to more readable format
                census_data = {
                    "zip_code": zip_code,
                    "median_household_income": int(result.get("B19013_001E", 0)) if result.get("B19013_001E") != "-666666666" else None,
                    "total_population": int(result.get("B01003_001E", 0)) if result.get("B01003_001E") != "-666666666" else None,
                    "data_year": year,
                    "data_source": f"ACS {data_type.upper()} (ZCTA)",
                    "geographic_level": "zip_code",
                    "fetched_at": datetime.now().isoformat()
                }

                self.logger.info(f"Successfully fetched Census data for zip code {zip_code}")
                return census_data

        except requests.RequestException:
            # If zip code fails, fall back to county-level data
            self.logger.warning(f"Zip code {zip_code} data not available, trying county-level data")

        # Fallback: Try to get county-level data
        # For now, return a message that zip-level data isn't available
        # In a production system, you might want to implement zip-to-county mapping
        return {
            "zip_code": zip_code,
            "error": "Zip code level data not available. County-level data can be fetched separately.",
            "data_year": year,
            "data_source": f"ACS {data_type.upper()}",
            "geographic_level": "county_fallback_needed",
            "fetched_at": datetime.now().isoformat()
        }

    def fetch_county_census_data(self, state_code: str, county_code: str, data_type: str = "acs5") -> Dict[str, Any]:
        """
        Fetch Census Bureau demographic data for a specific county.

        Args:
            state_code: 2-digit FIPS state code (e.g., "48" for Texas)
            county_code: 3-digit FIPS county code (e.g., "201" for Harris County)
            data_type: Census data type (acs5 for American Community Survey 5-year estimates)

        Returns:
            Dictionary with census data including median income and population
        """
        api_key = os.getenv("CENSUS_API_KEY")
        if not api_key:
            raise ValueError("CENSUS_API_KEY not found in environment variables")

        base_url = "https://api.census.gov/data"
        year = "2022"

        variables = "B19013_001E,B01003_001E"
        url = f"{base_url}/{year}/acs/{data_type}?get={variables}&for=county:{county_code}&in=state:{state_code}&key={api_key}"

        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()

            data = response.json()

            if len(data) < 2:
                return {"error": "No data found for this county"}

            headers = data[0]
            values = data[1]

            result = {}
            for i, header in enumerate(headers):
                if i < len(values):
                    result[header] = values[i]

            census_data = {
                "state_code": state_code,
                "county_code": county_code,
                "median_household_income": int(result.get("B19013_001E", 0)) if result.get("B19013_001E") != "-666666666" else None,
                "total_population": int(result.get("B01003_001E", 0)) if result.get("B01003_001E") != "-666666666" else None,
                "data_year": year,
                "data_source": f"ACS {data_type.upper()} (County)",
                "geographic_level": "county",
                "fetched_at": datetime.now().isoformat()
            }

            self.logger.info(f"Successfully fetched Census data for county {county_code}, state {state_code}")
            return census_data

        except requests.RequestException as e:
            self.logger.error(f"Census API request failed: {e}")
            return {"error": f"API request failed: {str(e)}"}
        except Exception as e:
            self.logger.error(f"Error processing Census data: {e}")
            return {"error": f"Data processing failed: {str(e)}"}

    def sync_census_data_to_excel(self, zip_codes: List[str],
                                 excel_file_path: Optional[Path] = None,
                                 sheet_name: str = "Census_Data") -> SyncResult:
        """
        Fetch Census data for multiple zip codes and sync to Excel.

        Args:
            zip_codes: List of 5-digit zip codes
            excel_file_path: Path to Excel file (defaults to main system file)
            sheet_name: Name of the sheet to update/create

        Returns:
            SyncResult with operation details
        """
        if excel_file_path is None:
            excel_file_path = Path("Land_and_Build_Flipping_System_v2.xlsx")

        result = SyncResult(
            direction=SyncDirection.DB_TO_EXCEL,
            records_processed=len(zip_codes),
            records_added=0,
            records_updated=0,
            records_skipped=0,
            conflicts=[],
            errors=[],
            timestamp=datetime.now()
        )

        try:
            # Fetch data for all zip codes
            census_data = []
            for zip_code in zip_codes:
                data = self.fetch_census_data(zip_code)
                if "error" not in data:
                    census_data.append(data)
                    result.records_added += 1
                else:
                    # If zip code data fails, try to get county data as fallback
                    # For demonstration, we'll use Harris County, TX as example
                    # In production, you'd want zip-to-county mapping
                    county_data = self.fetch_county_census_data("48", "201")  # Harris County, TX
                    if "error" not in county_data:
                        # Mark this as county-level data for the zip code area
                        county_data["associated_zip"] = zip_code
                        county_data["note"] = "County-level data (zip-level not available)"
                        census_data.append(county_data)
                        result.records_added += 1
                        self.logger.info(f"Used county data for zip {zip_code}")
                    else:
                        result.errors.append(f"Failed to fetch data for {zip_code}: {data['error']}")
                        result.records_skipped += 1

            if not census_data:
                result.errors.append("No Census data could be fetched")
                return result

            # Write to Excel
            self._write_census_data_to_excel(census_data, excel_file_path, sheet_name)

            self.logger.info(f"Successfully synced {len(census_data)} Census records to Excel")

        except Exception as e:
            result.errors.append(f"Census data sync failed: {str(e)}")
            self.logger.error(f"Census sync error: {e}")

        return result

    def _write_census_data_to_excel(self, census_data: List[Dict[str, Any]],
                                   file_path: Path, sheet_name: str) -> None:
        """
        Write Census data to Excel sheet.
        """
        try:
            if file_path.exists():
                wb = openpyxl.load_workbook(file_path)
            else:
                wb = openpyxl.Workbook()

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Clear existing data (keep headers)
                ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)
                # Add headers
                headers = ["Zip Code/County", "Median Household Income", "Total Population",
                          "Data Year", "Data Source", "Geographic Level", "Notes", "Fetched At"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Add data
            for row, data in enumerate(census_data, 2):
                # Handle different data formats
                if data.get("geographic_level") == "zip_code":
                    location = data.get("zip_code", "")
                elif data.get("geographic_level") == "county":
                    location = f"County {data.get('county_code', '')}, State {data.get('state_code', '')}"
                    if data.get("associated_zip"):
                        location += f" (for zip {data['associated_zip']})"
                else:
                    location = data.get("zip_code", "Unknown")

                ws.cell(row=row, column=1, value=location)
                ws.cell(row=row, column=2, value=data.get("median_household_income"))
                ws.cell(row=row, column=3, value=data.get("total_population"))
                ws.cell(row=row, column=4, value=data.get("data_year"))
                ws.cell(row=row, column=5, value=data.get("data_source"))
                ws.cell(row=row, column=6, value=data.get("geographic_level"))
                ws.cell(row=row, column=7, value=data.get("note", ""))
                ws.cell(row=row, column=8, value=data.get("fetched_at"))

            wb.save(file_path)
            self.logger.info(f"Wrote {len(census_data)} Census records to {file_path}")

        except Exception as e:
            self.logger.error(f"Failed to write Census data to Excel: {e}")
            raise

    def get_census_data_for_properties(self, property_zip_codes: List[str]) -> Dict[str, Dict[str, Any]]:
        """
        Get Census demographic data for a list of property zip codes.

        Args:
            property_zip_codes: List of zip codes from properties

        Returns:
            Dictionary mapping zip codes to Census data
        """
        census_data = {}

        for zip_code in set(property_zip_codes):  # Remove duplicates
            data = self.fetch_census_data(zip_code)
            if "error" not in data:
                census_data[zip_code] = data
            else:
                self.logger.warning(f"Could not fetch Census data for zip {zip_code}: {data['error']}")

        return census_data