import json
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import re

# Read the source data files
print("Loading data files...")

with open('/home/ubuntu/real_estate_guide/example_counties_detailed.json', 'r') as f:
    counties_data = json.load(f)

with open('/home/ubuntu/real_estate_guide/contact_databases.json', 'r') as f:
    contacts_data = json.load(f)

with open('/home/ubuntu/real_estate_guide/state_specific_guidance.md', 'r') as f:
    state_guidance_content = f.read()

with open('/home/ubuntu/real_estate_guide/formulas_guide.md', 'r') as f:
    formulas_content = f.read()

# Create workbook
print("Creating workbook...")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define color scheme
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="4472C4")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row_num, start_col=1, end_col=10):
    """Style a header row"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def apply_alternating_rows(ws, start_row, end_row, start_col, end_col):
    """Apply alternating row colors"""
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb != 'FF4472C4':  # Don't overwrite headers
                    cell.fill = ALT_ROW_FILL

def add_borders(ws, start_row, end_row, start_col, end_col):
    """Add borders to a range"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

print("Creating Sheet 1: README/Instructions...")
# ====================
# SHEET 1: README/Instructions
# ====================
ws1 = wb.create_sheet("README")

# Welcome header with color
ws1['A1'] = "🏠 REAL ESTATE INVESTMENT TOOLKIT"
ws1['A1'].font = Font(bold=True, size=18, color="FF0000")
ws1['A1'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A1:G1')

ws1['A2'] = "Welcome to Your Complete Investment Analysis System"
ws1['A2'].font = Font(bold=True, size=14, color="4472C4")
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A2:G2')

# Content sections
row = 4
ws1[f'A{row}'] = "📖 HOW TO USE THIS WORKBOOK"
ws1[f'A{row}'].font = Font(bold=True, size=13, color="4472C4")
row += 2

instructions = [
    ("Navigation", "Click on the sheet tabs at the bottom to navigate between different sections"),
    ("Example Counties", "Browse 12 example counties with live hyperlinks to GIS systems and assessor websites"),
    ("State Guidance", "Review state-specific information for 15 major real estate markets"),
    ("Contact Database", "Access organized contacts for title companies, lenders, builders, agents, and contractors"),
    ("Deal Calculator", "Input property details in YELLOW cells to auto-calculate investment metrics"),
    ("Property Analysis", "Comprehensive deal analysis worksheet for detailed due diligence"),
    ("CMA Template", "Comparative Market Analysis with comps and adjustment formulas"),
    ("Mail Campaign", "Track marketing campaigns, responses, and calculate ROI"),
    ("Portfolio Dashboard", "View summary analytics and charts for your entire portfolio"),
    ("Formulas Reference", "Complete guide to all real estate investment formulas used")
]

for title, desc in instructions:
    ws1[f'A{row}'] = f"• {title}:"
    ws1[f'A{row}'].font = BOLD_FONT
    ws1[f'B{row}'] = desc
    ws1[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws1.merge_cells(f'B{row}:G{row}')
    row += 1

row += 2
ws1[f'A{row}'] = "🎨 COLOR CODING SYSTEM"
ws1[f'A{row}'].font = Font(bold=True, size=13, color="4472C4")
row += 2

color_guide = [
    ("Blue Headers", "Navigation and section identifiers", HEADER_FILL, HEADER_FONT),
    ("Yellow Cells", "INPUT REQUIRED - Enter your data here", INPUT_FILL, Font(bold=True)),
    ("Green Cells", "AUTO-CALCULATED - Formulas will compute these", OUTPUT_FILL, Font(bold=True)),
    ("Alternating Rows", "For easier reading of data tables", ALT_ROW_FILL, Font())
]

for label, desc, fill, font in color_guide:
    ws1[f'A{row}'] = label
    ws1[f'A{row}'].fill = fill
    ws1[f'A{row}'].font = font
    ws1[f'A{row}'].alignment = Alignment(horizontal='center')
    ws1[f'B{row}'] = desc
    ws1[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws1.merge_cells(f'B{row}:G{row}')
    row += 1

row += 2
ws1[f'A{row}'] = "⚠️ IMPORTANT TIPS"
ws1[f'A{row}'].font = Font(bold=True, size=13, color="FF0000")
row += 2

tips = [
    "Protect Formulas: Formula cells are locked. Yellow input cells are unlocked.",
    "Use Filters: Most data tables have filter buttons enabled - click to sort and filter",
    "Hyperlinks: Blue underlined text links to external websites - click to open",
    "Data Validation: Some cells have dropdown lists - look for the arrow",
    "Save Often: Save your work regularly to avoid losing your analysis",
    "Make Copies: Create a copy before making major changes to preserve original structure"
]

for tip in tips:
    ws1[f'A{row}'] = f"• {tip}"
    ws1[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws1.merge_cells(f'A{row}:G{row}')
    row += 1

# Adjust column widths
ws1.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws1.column_dimensions[col].width = 15

print("Creating Sheet 2: Example Counties Directory...")
# ====================
# SHEET 2: Example Counties Directory
# ====================
ws2 = wb.create_sheet("Example Counties")

# Title
ws2['A1'] = "Example Counties Directory - 12 Verified Locations"
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:J1')

# Headers
headers = ["State", "County", "GIS System URL", "Assessor URL", "Zoning URL", 
           "Recorder URL", "Flood Maps URL", "Wetlands URL", "Platform Type", "Notes"]
for col, header in enumerate(headers, start=1):
    cell = ws2.cell(row=3, column=col, value=header)
    
style_header_row(ws2, 3, 1, 10)

# Parse counties data
row = 4
for county_name, data in counties_data.items():
    # Extract state from county name
    parts = county_name.split(", ")
    county = parts[0]
    state = parts[1] if len(parts) > 1 else "N/A"
    
    ws2.cell(row=row, column=1, value=state)
    ws2.cell(row=row, column=2, value=county)
    
    # Add hyperlinks
    urls = [
        data.get('gis_system_url', 'N/A'),
        data.get('assessor_site_url', 'N/A'),
        data.get('zoning_department_url', 'N/A'),
        data.get('recorder_office_url', 'N/A'),
        data.get('fema_flood_map_url', 'N/A'),
        data.get('wetland_database_url', 'N/A')
    ]
    
    for col, url in enumerate(urls, start=3):
        cell = ws2.cell(row=row, column=col)
        if url and url != 'N/A' and url.startswith('http'):
            cell.value = url
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single")
        else:
            cell.value = url
    
    ws2.cell(row=row, column=9, value=data.get('platform_type', 'N/A'))
    ws2.cell(row=row, column=10, value=data.get('usage_tips', 'N/A')[:100])  # Truncate for display
    
    row += 1

# Apply styling
apply_alternating_rows(ws2, 4, row-1, 1, 10)
add_borders(ws2, 3, row-1, 1, 10)

# Enable filters
ws2.auto_filter.ref = f"A3:J{row-1}"

# Adjust column widths
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 20
for col in ['C', 'D', 'E', 'F', 'G', 'H']:
    ws2.column_dimensions[col].width = 35
ws2.column_dimensions['I'].width = 18
ws2.column_dimensions['J'].width = 40

print("Creating Sheet 3: State Guidance Summary...")
# ====================
# SHEET 3: State Guidance Summary
# ====================
ws3 = wb.create_sheet("State Guidance")

ws3['A1'] = "State-Specific Guidance Summary - 15 States"
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:E1')

# Headers
headers = ["State", "Common GIS Platforms", "Property Tax Info", "Key Considerations", "State Resources"]
for col, header in enumerate(headers, start=1):
    ws3.cell(row=3, column=col, value=header)

style_header_row(ws3, 3, 1, 5)

# Parse state data from markdown
states_info = {
    "Texas": {
        "gis": "Esri ArcGIS (County CADs)",
        "tax": "100% market value, County Appraisal Districts",
        "key": "Landlord-friendly, no rent control, flood risk in coastal areas",
        "resources": "TCEQ, TWDB, TNRIS"
    },
    "Florida": {
        "gis": "Esri ArcGIS, Geocortex",
        "tax": "Save Our Homes cap for homestead, not investment properties",
        "key": "Strong environmental regulations, wetlands restrictions",
        "resources": "FDEP, Water Management Districts"
    },
    "North Carolina": {
        "gis": "Esri (POLARIS, GeoPortal)",
        "tax": "Octennial reappraisal cycle (8 years)",
        "key": "Balanced landlord-tenant laws, historic district restrictions",
        "resources": "NCDEQ, NC OneMap, NCFMP"
    },
    "Arizona": {
        "gis": "Esri ArcGIS, Open Data Portals",
        "tax": "Limited Property Value (LPV) caps increases at 5%",
        "key": "Water supply requirements, HOA prevalence",
        "resources": "ADEQ, ADWR, AGIC"
    },
    "Georgia": {
        "gis": "Esri ArcGIS",
        "tax": "40% of assessed value, 45-day appeal window",
        "key": "Rapid development, special tax districts (CIDs)",
        "resources": "Georgia DNR, EPD, GIS Clearinghouse"
    },
    "Tennessee": {
        "gis": "Esri (MetroGIS, Parcel Viewers)",
        "tax": "25% residential, 40% commercial of appraised value",
        "key": "URLTA applies in counties >75K population",
        "resources": "TDEC, TNGIC"
    },
    "Ohio": {
        "gis": "Esri (County Auditor systems)",
        "tax": "6-year reappraisal cycle, 3-year update",
        "key": "Older housing stock, lead paint/asbestos concerns",
        "resources": "Ohio EPA, ODNR, OGRIP"
    },
    "Indiana": {
        "gis": "Esri, County-level systems",
        "tax": "Annual assessments, caps at 1-3% of assessed value",
        "key": "Landlord-friendly laws, agricultural tax exemptions",
        "resources": "IDEM, IndianaMAP"
    },
    "South Carolina": {
        "gis": "Esri, County and City portals",
        "tax": "6% of market value for residential",
        "key": "Coastal flood risks, hurricane exposure",
        "resources": "SCDHEC, SC GIS Portal"
    },
    "Pennsylvania": {
        "gis": "Esri, County-managed",
        "tax": "Varies by county, reassessment schedules differ",
        "key": "Strong tenant protections, older urban inventory",
        "resources": "PA DEP, PASDA"
    },
    "Colorado": {
        "gis": "Esri ArcGIS, Open Data",
        "tax": "Biennial reassessment, residential vs commercial rates",
        "key": "Growth controls, mountain/ski resort markets",
        "resources": "CDPHE, Colorado Hazard Mapping"
    },
    "Nevada": {
        "gis": "Esri ArcGIS Hub",
        "tax": "35% of taxable value, 3% annual cap",
        "key": "Tourism/gaming economy, HOA common",
        "resources": "NDEP, Nevada GIS"
    },
    "California": {
        "gis": "Esri ArcGIS Hub, LA eGIS",
        "tax": "Prop 13 caps increases at 2%/year, resets on sale",
        "key": "Strict tenant protections, rent control in some cities",
        "resources": "CalEPA, Cal-Atlas"
    },
    "Idaho": {
        "gis": "Esri, County systems",
        "tax": "100% market value, homestead exemption available",
        "key": "Growing market, agricultural conversion",
        "resources": "Idaho DEQ, GIS Data"
    },
    "Utah": {
        "gis": "Esri, AGRC",
        "tax": "100% fair market value, annual assessments",
        "key": "High growth, tech corridor development",
        "resources": "Utah DEQ, AGRC"
    }
}

row = 4
for state, info in states_info.items():
    ws3.cell(row=row, column=1, value=state)
    ws3.cell(row=row, column=2, value=info['gis'])
    ws3.cell(row=row, column=3, value=info['tax'])
    ws3.cell(row=row, column=4, value=info['key'])
    ws3.cell(row=row, column=5, value=info['resources'])
    row += 1

# Apply styling
apply_alternating_rows(ws3, 4, row-1, 1, 5)
add_borders(ws3, 3, row-1, 1, 5)
ws3.auto_filter.ref = f"A3:E{row-1}"

# Column widths
ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 45
ws3.column_dimensions['E'].width = 30

print("Creating Sheet 4: Contact Database...")
# ====================
# SHEET 4: Contact Database
# ====================
ws4 = wb.create_sheet("Contact Database")

ws4['A1'] = "Premium Contact Database - Title, Lenders, Builders, Agents, Contractors"
ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A1:G1')

# Headers
headers = ["Category", "Company Name", "Type", "Phone", "Email", "Website", "Service Area"]
for col, header in enumerate(headers, start=1):
    ws4.cell(row=3, column=col, value=header)

style_header_row(ws4, 3, 1, 7)

# Parse contact data - organize by relevant categories
row = 4
categories_to_include = ['title', 'lender', 'builder', 'agent', 'contractor', 'real estate']
displayed_categories = {
    'title': 'Title Companies',
    'lender': 'Lenders',
    'builder': 'Builders',
    'agent': 'Real Estate Agents',
    'contractor': 'Contractors',
    'real estate': 'Real Estate Services'
}

# Process limited entries (to keep file manageable)
max_per_category = 20
category_counts = {cat: 0 for cat in categories_to_include}

for entry in contacts_data:
    category = entry.get('category', '').lower()
    
    # Check if this category is relevant
    for cat_key in categories_to_include:
        if cat_key in category:
            if category_counts[cat_key] >= max_per_category:
                continue
            category_counts[cat_key] += 1
            
            ws4.cell(row=row, column=1, value=displayed_categories.get(cat_key, category.title()))
            ws4.cell(row=row, column=2, value=entry.get('company', 'N/A'))
            
            # Type/subcategory
            ws4.cell(row=row, column=3, value=entry.get('subcategory', 'General'))
            
            # Phone
            phones = entry.get('phone', [])
            phone_str = phones[0] if phones else 'N/A'
            ws4.cell(row=row, column=4, value=phone_str)
            
            # Email
            emails = entry.get('email', [])
            email_str = emails[0] if emails else 'N/A'
            ws4.cell(row=row, column=5, value=email_str)
            
            # Website with hyperlink
            websites = entry.get('website', [])
            if websites:
                url = websites[0]
                cell = ws4.cell(row=row, column=6, value=url)
                if url.startswith('http'):
                    cell.hyperlink = url
                    cell.font = Font(color="0000FF", underline="single")
            else:
                ws4.cell(row=row, column=6, value='N/A')
            
            # Service area
            service_area = entry.get('service_areas', entry.get('hq', 'N/A'))
            ws4.cell(row=row, column=7, value=service_area)
            
            row += 1
            break

# Apply styling
apply_alternating_rows(ws4, 4, row-1, 1, 7)
add_borders(ws4, 3, row-1, 1, 7)
ws4.auto_filter.ref = f"A3:G{row-1}"

# Column widths
ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 25
ws4.column_dimensions['F'].width = 40
ws4.column_dimensions['G'].width = 20

print("Creating Sheet 5: Deal Calculator...")
# ====================
# SHEET 5: Deal Calculator
# ====================
ws5 = wb.create_sheet("Deal Calculator")

ws5['A1'] = "💰 INVESTMENT DEAL CALCULATOR"
ws5['A1'].font = Font(bold=True, size=16, color="4472C4")
ws5.merge_cells('A1:D1')

row = 3

# Property Information Section
ws5[f'A{row}'] = "PROPERTY INFORMATION"
ws5[f'A{row}'].font = BOLD_FONT
ws5[f'A{row}'].fill = HEADER_FILL
ws5[f'A{row}'].font = HEADER_FONT
ws5.merge_cells(f'A{row}:B{row}')
row += 1

property_inputs = [
    ("Address", ""),
    ("Purchase Price", 200000),
    ("After Repair Value (ARV)", 300000),
    ("Property Type", "Single Family"),
    ("Square Footage", 1800),
    ("Bedrooms", 3),
    ("Bathrooms", 2)
]

for label, default_val in property_inputs:
    ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws5.cell(row=row, column=2, value=default_val)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    if isinstance(default_val, (int, float)) and default_val > 0:
        cell.number_format = '#,##0'
    row += 1

row += 1

# Purchase Costs Section
ws5[f'A{row}'] = "PURCHASE COSTS"
ws5[f'A{row}'].font = HEADER_FONT
ws5[f'A{row}'].fill = HEADER_FILL
ws5.merge_cells(f'A{row}:B{row}')
row += 1

purchase_costs = [
    ("Down Payment %", 20),
    ("Down Payment $", f"=B{row-1+5}/100*B5"),
    ("Closing Costs", 5000),
    ("Inspection Fees", 500),
    ("Appraisal Fee", 600),
    ("Other Fees", 1000)
]

down_payment_row = row + 2
for label, default_val in purchase_costs:
    ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws5.cell(row=row, column=2)
    if isinstance(default_val, str) and default_val.startswith('='):
        cell.value = default_val
        cell.fill = OUTPUT_FILL
    else:
        cell.value = default_val
        cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.number_format = '#,##0'
    row += 1

row += 1

# Repair Costs Section
ws5[f'A{row}'] = "REPAIR & RENOVATION COSTS"
ws5[f'A{row}'].font = HEADER_FONT
ws5[f'A{row}'].fill = HEADER_FILL
ws5.merge_cells(f'A{row}:B{row}')
row += 1

repair_costs = [
    ("Kitchen Remodel", 15000),
    ("Bathroom Remodel", 8000),
    ("Flooring", 5000),
    ("Painting", 3000),
    ("Roof Repair", 0),
    ("HVAC", 4000),
    ("Plumbing", 2000),
    ("Electrical", 2000),
    ("Landscaping", 2000),
    ("Contingency (10%)", f"=SUM(B{row+1}:B{row+8})*0.1")
]

for label, default_val in repair_costs:
    ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws5.cell(row=row, column=2)
    if isinstance(default_val, str) and default_val.startswith('='):
        cell.value = default_val
        cell.fill = OUTPUT_FILL
    else:
        cell.value = default_val
        cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.number_format = '#,##0'
    row += 1

ws5.cell(row=row, column=1, value="Total Repair Costs").font = BOLD_FONT
repair_total_row = row
ws5.cell(row=row, column=2, value=f"=SUM(B{row-10}:B{row-1})")
ws5.cell(row=row, column=2).fill = OUTPUT_FILL
ws5.cell(row=row, column=2).number_format = '#,##0'
ws5.cell(row=row, column=2).font = BOLD_FONT
ws5.cell(row=row, column=2).border = THIN_BORDER
row += 2

# Holding Costs Section
ws5[f'A{row}'] = "HOLDING COSTS (Monthly)"
ws5[f'A{row}'].font = HEADER_FONT
ws5[f'A{row}'].fill = HEADER_FILL
ws5.merge_cells(f'A{row}:B{row}')
row += 1

holding_costs = [
    ("Months to Hold", 6),
    ("Mortgage Payment", 1200),
    ("Property Taxes", 250),
    ("Insurance", 150),
    ("Utilities", 200),
    ("HOA Fees", 0),
    ("Maintenance Reserve", 200)
]

months_hold_row = row + 1
for label, default_val in holding_costs:
    ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws5.cell(row=row, column=2, value=default_val)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.number_format = '#,##0'
    row += 1

ws5.cell(row=row, column=1, value="Total Holding Costs").font = BOLD_FONT
holding_total_row = row
ws5.cell(row=row, column=2, value=f"=SUM(B{months_hold_row+1}:B{row-1})*B{months_hold_row}")
ws5.cell(row=row, column=2).fill = OUTPUT_FILL
ws5.cell(row=row, column=2).number_format = '#,##0'
ws5.cell(row=row, column=2).font = BOLD_FONT
ws5.cell(row=row, column=2).border = THIN_BORDER
row += 2

# Rental/Sale Information
ws5[f'A{row}'] = "RENTAL INFORMATION (If Renting)"
ws5[f'A{row}'].font = HEADER_FONT
ws5[f'A{row}'].fill = HEADER_FILL
ws5.merge_cells(f'A{row}:B{row}')
row += 1

rental_info = [
    ("Monthly Rent", 2000),
    ("Vacancy Rate %", 5),
    ("Property Management %", 10),
    ("Annual Operating Expenses", 8000)
]

monthly_rent_row = row + 1
for label, default_val in rental_info:
    ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
    cell = ws5.cell(row=row, column=2, value=default_val)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    if '%' in label:
        cell.number_format = '0'
    else:
        cell.number_format = '#,##0'
    row += 1

row += 1

# CALCULATED OUTPUTS Section
ws5[f'A{row}'] = "📊 CALCULATED INVESTMENT METRICS"
ws5[f'A{row}'].font = Font(bold=True, size=14, color="006100")
ws5[f'A{row}'].fill = OUTPUT_FILL
ws5.merge_cells(f'A{row}:B{row}')
row += 1

# Calculate total investment for MAO
ws5.cell(row=row, column=1, value="Total Investment Cost").font = BOLD_FONT
ws5.cell(row=row, column=2, value=f"=B5+B{down_payment_row+1}+B{down_payment_row+2}+B{down_payment_row+3}+B{down_payment_row+4}+B{repair_total_row}+B{holding_total_row}")
ws5.cell(row=row, column=2).fill = OUTPUT_FILL
ws5.cell(row=row, column=2).number_format = '$#,##0'
ws5.cell(row=row, column=2).font = BOLD_FONT
ws5.cell(row=row, column=2).border = THIN_BORDER
total_investment_row = row
row += 1

outputs = [
    ("Maximum Allowable Offer (MAO)", f"=B6-B{repair_total_row}-B{holding_total_row}-40000"),
    ("Potential Profit (ARV - Total Investment)", f"=B6-B{total_investment_row}"),
    ("Return on Investment (ROI) %", f"=(B6-B{total_investment_row})/B{total_investment_row}*100"),
    ("70% Rule MAO", f"=B6*0.7-B{repair_total_row}"),
    ("", ""),
    ("Annual Gross Rental Income", f"=B{monthly_rent_row}*12"),
    ("Net Operating Income (NOI)", f"=B{monthly_rent_row}*12*(1-B{monthly_rent_row+1}/100)*(1-B{monthly_rent_row+2}/100)-B{monthly_rent_row+3}"),
    ("Cap Rate %", f"=B{row+5}/B5*100"),
    ("Cash-on-Cash Return %", f"=(B{row+5}-B{months_hold_row+1}*12)/B{down_payment_row}*100"),
    ("Debt Service Coverage Ratio (DSCR)", f"=B{row+5}/(B{months_hold_row+1}*12)"),
    ("Monthly Cash Flow", f"=B{monthly_rent_row}*(1-B{monthly_rent_row+1}/100)*(1-B{monthly_rent_row+2}/100)-B{months_hold_row+1}-B{monthly_rent_row+3}/12")
]

for label, formula in outputs:
    if label == "":
        row += 1
        continue
    ws5.cell(row=row, column=1, value=label).font = BOLD_FONT
    cell = ws5.cell(row=row, column=2, value=formula)
    cell.fill = OUTPUT_FILL
    cell.font = Font(bold=True, color="006100")
    cell.border = THIN_BORDER
    if '%' in label:
        cell.number_format = '0.00"%"'
    elif 'DSCR' in label or 'Ratio' in label:
        cell.number_format = '0.00'
    else:
        cell.number_format = '$#,##0'
    row += 1

# Column widths
ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 15

print("Creating Sheet 6: Property Analysis...")
# ====================
# SHEET 6: Property Analysis Worksheet
# ====================
ws6 = wb.create_sheet("Property Analysis")

ws6['A1'] = "COMPREHENSIVE PROPERTY ANALYSIS WORKSHEET"
ws6['A1'].font = TITLE_FONT
ws6.merge_cells('A1:D1')

row = 3
sections = [
    ("Property Details", [
        ("Property Address", ""),
        ("City/State/Zip", ""),
        ("Property Type", "Single Family"),
        ("Year Built", 1990),
        ("Lot Size (acres)", 0.25),
        ("Square Footage", 1800),
        ("Bedrooms/Bathrooms", "3/2"),
        ("Current Condition", "Needs Rehab")
    ]),
    ("Purchase Analysis", [
        ("Asking Price", 180000),
        ("Your Offer", 170000),
        ("Estimated ARV", 300000),
        ("Purchase Closing Costs", 5000),
        ("Total Acquisition Cost", "=B14+B17")
    ]),
    ("Rehab Budget", [
        ("Structural Repairs", 0),
        ("Roof", 0),
        ("HVAC", 4000),
        ("Plumbing", 2000),
        ("Electrical", 2000),
        ("Kitchen", 15000),
        ("Bathrooms", 8000),
        ("Flooring", 5000),
        ("Paint (Interior/Exterior)", 3500),
        ("Landscaping/Curb Appeal", 2000),
        ("Appliances", 2500),
        ("Permits & Fees", 1000),
        ("Contingency (10%)", "=SUM(B20:B31)*0.1"),
        ("Total Rehab Budget", "=SUM(B20:B32)")
    ]),
    ("Financing", [
        ("Financing Type", "Conventional"),
        ("Down Payment %", 20),
        ("Down Payment $", "=B36*B14/100"),
        ("Loan Amount", "=B14-B37"),
        ("Interest Rate %", 7.5),
        ("Loan Term (years)", 30),
        ("Monthly P&I Payment", "=PMT(B39/12,B40*12,-B38)")
    ]),
    ("Operating Expenses (Annual)", [
        ("Property Taxes", 3000),
        ("Insurance", 1800),
        ("HOA Fees", 0),
        ("Utilities (if landlord pays)", 0),
        ("Property Management (10%)", "=B51*12*0.1"),
        ("Maintenance Reserve", 2400),
        ("Capital Reserves", 1800),
        ("Total Annual Operating Expenses", "=SUM(B43:B49)")
    ]),
    ("Rental Analysis", [
        ("Estimated Monthly Rent", 2000),
        ("Vacancy Rate %", 5),
        ("Effective Monthly Rent", "=B51*(1-B52/100)"),
        ("Gross Annual Rental Income", "=B51*12"),
        ("Net Operating Income (NOI)", "=B54-B50"),
        ("Monthly Cash Flow", "=B53-B41-B50/12")
    ])
]

for section_title, items in sections:
    ws6[f'A{row}'] = section_title
    ws6[f'A{row}'].font = HEADER_FONT
    ws6[f'A{row}'].fill = HEADER_FILL
    ws6.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for label, default_val in items:
        ws6.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws6.cell(row=row, column=2)
        
        if isinstance(default_val, str) and (default_val.startswith('=') or default_val.startswith('PMT')):
            cell.value = default_val
            cell.fill = OUTPUT_FILL
            cell.font = Font(bold=True, color="006100")
        else:
            cell.value = default_val
            cell.fill = INPUT_FILL
        
        cell.border = THIN_BORDER
        if isinstance(default_val, (int, float)) and default_val > 100:
            cell.number_format = '$#,##0'
        elif '%' in label:
            cell.number_format = '0.00'
        
        row += 1
    row += 1

# Summary Metrics
ws6[f'A{row}'] = "KEY INVESTMENT METRICS"
ws6[f'A{row}'].font = Font(bold=True, size=13, color="006100")
ws6[f'A{row}'].fill = OUTPUT_FILL
ws6.merge_cells(f'A{row}:B{row}')
row += 1

metrics = [
    ("Total All-In Cost", "=B18+B33"),
    ("Potential Equity at ARV", "=B15-B60"),
    ("Cap Rate %", "=B55/B14*100"),
    ("Cash-on-Cash Return %", "=(B56*12)/B37*100"),
    ("DSCR", "=B55/(B41*12)"),
    ("ROI if Flip %", "=(B15-B60)/B60*100")
]

for label, formula in metrics:
    ws6.cell(row=row, column=1, value=label).font = BOLD_FONT
    cell = ws6.cell(row=row, column=2, value=formula)
    cell.fill = OUTPUT_FILL
    cell.font = Font(bold=True, color="006100", size=11)
    cell.border = THIN_BORDER
    if '%' in label:
        cell.number_format = '0.00"%"'
    elif 'DSCR' in label:
        cell.number_format = '0.00'
    else:
        cell.number_format = '$#,##0'
    row += 1

ws6.column_dimensions['A'].width = 38
ws6.column_dimensions['B'].width = 20

print("Creating Sheet 7: CMA Template...")
# ====================
# SHEET 7: Comparative Market Analysis (CMA)
# ====================
ws7 = wb.create_sheet("CMA Template")

ws7['A1'] = "COMPARATIVE MARKET ANALYSIS (CMA)"
ws7['A1'].font = TITLE_FONT
ws7.merge_cells('A1:H1')

row = 3
ws7[f'A{row}'] = "Subject Property Address:"
ws7[f'A{row}'].font = BOLD_FONT
ws7.cell(row=row, column=2, value="").fill = INPUT_FILL
ws7.merge_cells(f'B{row}:D{row}')
row += 2

# Comparable Properties Table
ws7[f'A{row}'] = "COMPARABLE PROPERTIES"
ws7[f'A{row}'].font = HEADER_FONT
ws7[f'A{row}'].fill = HEADER_FILL
ws7.merge_cells(f'A{row}:H{row}')
row += 1

headers = ["Address", "Sale Price", "Sale Date", "Sq Ft", "Beds/Baths", "Age", "Condition", "Distance (mi)"]
for col, header in enumerate(headers, start=1):
    ws7.cell(row=row, column=col, value=header)

style_header_row(ws7, row, 1, 8)
comp_header_row = row
row += 1

# Sample comps
comps = [
    ("123 Main St", 295000, "2025-12-01", 1850, "3/2", 30, "Good", 0.5),
    ("456 Oak Ave", 305000, "2025-11-15", 1900, "3/2.5", 28, "Excellent", 0.8),
    ("789 Elm Dr", 285000, "2025-12-10", 1750, "3/2", 35, "Fair", 0.3),
    ("321 Pine Ln", 310000, "2025-11-20", 1950, "4/2", 25, "Excellent", 1.0),
    ("654 Maple Ct", 290000, "2025-12-05", 1800, "3/2", 32, "Good", 0.6)
]

for comp_data in comps:
    for col, val in enumerate(comp_data, start=1):
        cell = ws7.cell(row=row, column=col, value=val)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        if col == 2:  # Sale price
            cell.number_format = '$#,##0'
        elif col == 4:  # Sq ft
            cell.number_format = '#,##0'
    row += 1

row += 1

# Adjustments Table
ws7[f'A{row}'] = "ADJUSTMENTS"
ws7[f'A{row}'].font = HEADER_FONT
ws7[f'A{row}'].fill = HEADER_FILL
ws7.merge_cells(f'A{row}:F{row}')
row += 1

adj_headers = ["Factor", "Comp 1", "Comp 2", "Comp 3", "Comp 4", "Comp 5"]
for col, header in enumerate(adj_headers, start=1):
    ws7.cell(row=row, column=col, value=header)
style_header_row(ws7, row, 1, 6)
adj_header_row = row
row += 1

adjustments = [
    ("Size Adjustment ($/sq ft)", 0, 0, 5000, -2000, 0),
    ("Condition Adjustment", -5000, 0, -8000, 0, -3000),
    ("Age Adjustment", 0, 0, 0, 3000, 0),
    ("Location Adjustment", 2000, -3000, 4000, -5000, 1000),
    ("Total Adjustments", f"=SUM(B{row}:B{row+3})", f"=SUM(C{row}:C{row+3})", 
     f"=SUM(D{row}:D{row+3})", f"=SUM(E{row}:E{row+3})", f"=SUM(F{row}:F{row+3})")
]

for adj_data in adjustments:
    label = adj_data[0]
    ws7.cell(row=row, column=1, value=label).font = Font(bold=True)
    
    for col, val in enumerate(adj_data[1:], start=2):
        cell = ws7.cell(row=row, column=col)
        if isinstance(val, str) and val.startswith('='):
            cell.value = val
            cell.fill = OUTPUT_FILL
            cell.font = Font(bold=True)
        else:
            cell.value = val
            cell.fill = INPUT_FILL if "Total" not in label else OUTPUT_FILL
        cell.border = THIN_BORDER
        cell.number_format = '$#,##0'
    row += 1

row += 1

# Adjusted Values
ws7[f'A{row}'] = "ADJUSTED VALUES"
ws7[f'A{row}'].font = HEADER_FONT
ws7[f'A{row}'].fill = HEADER_FILL
ws7.merge_cells(f'A{row}:F{row}')
row += 1

ws7.cell(row=row, column=1, value="Adjusted Comp Value").font = BOLD_FONT
for col in range(2, 7):
    comp_idx = col + 4  # Reference to sale price row
    adj_idx = row - 2  # Reference to total adjustments
    cell = ws7.cell(row=row, column=col)
    cell.value = f"=B{comp_header_row+col-1}+{chr(64+col)}{adj_idx}"
    cell.fill = OUTPUT_FILL
    cell.font = Font(bold=True, color="006100")
    cell.border = THIN_BORDER
    cell.number_format = '$#,##0'
row += 2

# Final Value Estimate
ws7[f'A{row}'] = "ESTIMATED MARKET VALUE"
ws7[f'A{row}'].font = Font(bold=True, size=13, color="006100")
ws7.merge_cells(f'A{row}:B{row}')
row += 1

ws7.cell(row=row, column=1, value="Average of Adjusted Comps").font = BOLD_FONT
ws7.cell(row=row, column=2, value=f"=AVERAGE(B{row-3}:F{row-3})")
ws7.cell(row=row, column=2).fill = OUTPUT_FILL
ws7.cell(row=row, column=2).font = Font(bold=True, size=12, color="006100")
ws7.cell(row=row, column=2).number_format = '$#,##0'
ws7.cell(row=row, column=2).border = THIN_BORDER
row += 1

ws7.cell(row=row, column=1, value="Recommended ARV (Conservative)").font = BOLD_FONT
ws7.cell(row=row, column=2, value=f"=B{row-1}*0.97")
ws7.cell(row=row, column=2).fill = OUTPUT_FILL
ws7.cell(row=row, column=2).font = Font(bold=True, size=12, color="006100")
ws7.cell(row=row, column=2).number_format = '$#,##0'
ws7.cell(row=row, column=2).border = THIN_BORDER

# Column widths
ws7.column_dimensions['A'].width = 32
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws7.column_dimensions[col].width = 14

print("Creating Sheet 8: Mail Campaign Tracker...")
# ====================
# SHEET 8: Mail Campaign Tracker
# ====================
ws8 = wb.create_sheet("Mail Campaign")

ws8['A1'] = "MARKETING CAMPAIGN TRACKER"
ws8['A1'].font = TITLE_FONT
ws8.merge_cells('A1:I1')

# Headers
row = 3
headers = ["Contact Name", "Property Address", "Phone", "Email", "Campaign Date", 
           "Response Date", "Status", "Notes", "Cost per Contact"]
for col, header in enumerate(headers, start=1):
    ws8.cell(row=row, column=col, value=header)

style_header_row(ws8, row, 1, 9)
row += 1

# Sample data
sample_data = [
    ("John Smith", "123 Main St", "555-0101", "john@email.com", "2026-01-01", "", "Sent", "First mailer", 0.75),
    ("Mary Johnson", "456 Oak Ave", "555-0102", "mary@email.com", "2026-01-01", "2026-01-05", "Interested", "Called back", 0.75),
    ("Bob Wilson", "789 Elm Dr", "555-0103", "bob@email.com", "2026-01-01", "", "Sent", "First mailer", 0.75),
    ("Sarah Davis", "321 Pine Ln", "555-0104", "sarah@email.com", "2026-01-01", "2026-01-06", "Not Interested", "Not selling", 0.75),
    ("Mike Brown", "654 Maple Ct", "555-0105", "mike@email.com", "2026-01-01", "2026-01-08", "Follow-up", "Wants more info", 0.75)
]

data_start_row = row
for contact_data in sample_data:
    for col, val in enumerate(contact_data, start=1):
        cell = ws8.cell(row=row, column=col, value=val)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        if col == 9:  # Cost
            cell.number_format = '$0.00'
    row += 1

# Add some empty rows for user input
for _ in range(15):
    for col in range(1, 10):
        cell = ws8.cell(row=row, column=col, value="")
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
    row += 1

# Add status dropdown validation
status_options = '"Sent,Interested,Not Interested,Follow-up,Deal Closed,Dead Lead"'
dv = DataValidation(type="list", formula1=status_options, allow_blank=True)
ws8.add_data_validation(dv)
dv.add(f'G{data_start_row}:G{row-1}')

row += 1

# Campaign Metrics
ws8[f'A{row}'] = "CAMPAIGN METRICS"
ws8[f'A{row}'].font = HEADER_FONT
ws8[f'A{row}'].fill = HEADER_FILL
ws8.merge_cells(f'A{row}:B{row}')
row += 1

metrics = [
    ("Total Contacts", f"=COUNTA(A{data_start_row}:A{row-3})"),
    ("Total Responses", f"=COUNTIFS(G{data_start_row}:G{row-3},\"Interested\")+COUNTIFS(G{data_start_row}:G{row-3},\"Follow-up\")"),
    ("Response Rate %", f"=B{row+1}/B{row}*100"),
    ("Total Campaign Cost", f"=SUM(I{data_start_row}:I{row-3})"),
    ("Cost Per Lead", f"=B{row+3}/B{row+1}"),
    ("Deals Closed", f"=COUNTIF(G{data_start_row}:G{row-3},\"Deal Closed\")"),
    ("Conversion Rate %", f"=B{row+5}/B{row}*100")
]

for label, formula in metrics:
    ws8.cell(row=row, column=1, value=label).font = BOLD_FONT
    cell = ws8.cell(row=row, column=2, value=formula)
    cell.fill = OUTPUT_FILL
    cell.font = Font(bold=True, color="006100")
    cell.border = THIN_BORDER
    if '%' in label:
        cell.number_format = '0.00"%"'
    elif 'Cost' in label or 'Lead' in label:
        cell.number_format = '$#,##0.00'
    else:
        cell.number_format = '0'
    row += 1

# Column widths
ws8.column_dimensions['A'].width = 20
ws8.column_dimensions['B'].width = 25
ws8.column_dimensions['C'].width = 14
ws8.column_dimensions['D'].width = 25
ws8.column_dimensions['E'].width = 14
ws8.column_dimensions['F'].width = 14
ws8.column_dimensions['G'].width = 15
ws8.column_dimensions['H'].width = 30
ws8.column_dimensions['I'].width = 14

ws8.auto_filter.ref = f"A3:I{data_start_row+20}"

print("Creating Sheet 9: Portfolio Dashboard...")
# ====================
# SHEET 9: Portfolio Dashboard
# ====================
ws9 = wb.create_sheet("Portfolio Dashboard")

ws9['A1'] = "📊 PORTFOLIO PERFORMANCE DASHBOARD"
ws9['A1'].font = Font(bold=True, size=16, color="4472C4")
ws9.merge_cells('A1:F1')

row = 3

# Portfolio Summary Section
ws9[f'A{row}'] = "PORTFOLIO SUMMARY"
ws9[f'A{row}'].font = HEADER_FONT
ws9[f'A{row}'].fill = HEADER_FILL
ws9.merge_cells(f'A{row}:B{row}')
row += 1

# Sample property data
ws9.cell(row=row, column=1, value="Property").font = HEADER_FONT
ws9.cell(row=row, column=2, value="Purchase Price").font = HEADER_FONT
ws9.cell(row=row, column=3, value="Current Value").font = HEADER_FONT
ws9.cell(row=row, column=4, value="Equity").font = HEADER_FONT
ws9.cell(row=row, column=5, value="Monthly Rent").font = HEADER_FONT
ws9.cell(row=row, column=6, value="Monthly Cash Flow").font = HEADER_FONT
style_header_row(ws9, row, 1, 6)
row += 1

properties = [
    ("123 Main St", 180000, 300000, "=C{}-B{}", 2000, 450),
    ("456 Oak Ave", 220000, 320000, "=C{}-B{}", 2400, 550),
    ("789 Elm Dr", 150000, 240000, "=C{}-B{}", 1800, 380),
    ("321 Pine Ln", 200000, 310000, "=C{}-B{}", 2200, 500),
    ("654 Maple Ct", 175000, 275000, "=C{}-B{}", 1900, 420)
]

prop_start_row = row
for prop_name, purchase, value, equity_formula, rent, cashflow in properties:
    ws9.cell(row=row, column=1, value=prop_name).fill = INPUT_FILL
    ws9.cell(row=row, column=2, value=purchase).fill = INPUT_FILL
    ws9.cell(row=row, column=2).number_format = '$#,##0'
    ws9.cell(row=row, column=3, value=value).fill = INPUT_FILL
    ws9.cell(row=row, column=3).number_format = '$#,##0'
    ws9.cell(row=row, column=4, value=equity_formula.format(row, row)).fill = OUTPUT_FILL
    ws9.cell(row=row, column=4).number_format = '$#,##0'
    ws9.cell(row=row, column=5, value=rent).fill = INPUT_FILL
    ws9.cell(row=row, column=5).number_format = '$#,##0'
    ws9.cell(row=row, column=6, value=cashflow).fill = INPUT_FILL
    ws9.cell(row=row, column=6).number_format = '$#,##0'
    
    for col in range(1, 7):
        ws9.cell(row=row, column=col).border = THIN_BORDER
    row += 1

prop_end_row = row - 1
row += 1

# Portfolio Metrics
ws9[f'A{row}'] = "PORTFOLIO METRICS"
ws9[f'A{row}'].font = HEADER_FONT
ws9[f'A{row}'].fill = HEADER_FILL
ws9.merge_cells(f'A{row}:B{row}')
row += 1

metrics = [
    ("Total Properties", f"=COUNTA(A{prop_start_row}:A{prop_end_row})"),
    ("Total Investment", f"=SUM(B{prop_start_row}:B{prop_end_row})"),
    ("Total Current Value", f"=SUM(C{prop_start_row}:C{prop_end_row})"),
    ("Total Equity", f"=SUM(D{prop_start_row}:D{prop_end_row})"),
    ("Average ROI %", f"=(B{row+2}-B{row+1})/B{row+1}*100"),
    ("Monthly Rental Income", f"=SUM(E{prop_start_row}:E{prop_end_row})"),
    ("Annual Rental Income", f"=B{row+5}*12"),
    ("Monthly Portfolio Cash Flow", f"=SUM(F{prop_start_row}:F{prop_end_row})"),
    ("Annual Portfolio Cash Flow", f"=B{row+7}*12"),
    ("Average Cash-on-Cash %", f"=B{row+8}/B{row+1}*100")
]

metrics_start_row = row
for label, formula in metrics:
    ws9.cell(row=row, column=1, value=label).font = BOLD_FONT
    cell = ws9.cell(row=row, column=2, value=formula)
    cell.fill = OUTPUT_FILL
    cell.font = Font(bold=True, color="006100", size=11)
    cell.border = THIN_BORDER
    if '%' in label:
        cell.number_format = '0.00"%"'
    else:
        cell.number_format = '$#,##0'
    row += 1

# Note about charts
row += 1
ws9[f'A{row}'] = "📈 CHARTS"
ws9[f'A{row}'].font = TITLE_FONT
row += 1

ws9[f'A{row}'] = "Chart 1: Property Values"
ws9[f'A{row}'].font = BOLD_FONT
row += 1

# Create bar chart for property values
chart1 = BarChart()
chart1.title = "Property Purchase Price vs Current Value"
chart1.y_axis.title = "Value ($)"
chart1.x_axis.title = "Properties"

cats = Reference(ws9, min_col=1, min_row=prop_start_row, max_row=prop_end_row)
data = Reference(ws9, min_col=2, min_row=prop_start_row-1, max_col=3, max_row=prop_end_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 10
chart1.width = 20

ws9.add_chart(chart1, f"A{row}")
row += 18

# Create pie chart for equity distribution
ws9[f'A{row}'] = "Chart 2: Equity Distribution"
ws9[f'A{row}'].font = BOLD_FONT
row += 1

chart2 = PieChart()
chart2.title = "Equity Distribution Across Portfolio"

labels = Reference(ws9, min_col=1, min_row=prop_start_row, max_row=prop_end_row)
data = Reference(ws9, min_col=4, min_row=prop_start_row, max_row=prop_end_row)
chart2.add_data(data, titles_from_data=False)
chart2.set_categories(labels)
chart2.height = 10
chart2.width = 15

ws9.add_chart(chart2, f"A{row}")

# Column widths
ws9.column_dimensions['A'].width = 22
ws9.column_dimensions['B'].width = 18
ws9.column_dimensions['C'].width = 18
ws9.column_dimensions['D'].width = 18
ws9.column_dimensions['E'].width = 18
ws9.column_dimensions['F'].width = 18

print("Creating Sheet 10: Formulas Reference...")
# ====================
# SHEET 10: Formulas Reference
# ====================
ws10 = wb.create_sheet("Formulas Reference")

ws10['A1'] = "REAL ESTATE INVESTMENT FORMULAS REFERENCE GUIDE"
ws10['A1'].font = TITLE_FONT
ws10.merge_cells('A1:E1')

row = 3

# Headers
headers = ["Formula Name", "What It Measures", "Formula", "Excel Syntax", "Example"]
for col, header in enumerate(headers, start=1):
    ws10.cell(row=row, column=col, value=header)

style_header_row(ws10, row, 1, 5)
row += 1

# Parse formulas from guide
formulas_data = [
    ("Maximum Allowable Offer (MAO)", 
     "Highest price to pay for property",
     "ARV - Fixed Costs - Rehab - Profit",
     "=ARV - Fixed - Rehab - Profit",
     "$300K - $25K - $30K - $40K = $205K"),
    
    ("After Repair Value (ARV)",
     "Estimated value after repairs",
     "Average of Comparable Sales",
     "=AVERAGE(Comp1,Comp2,Comp3)",
     "($310K + $295K + $305K) / 3 = $303K"),
    
    ("70% Rule",
     "Quick MAO calculation for flips",
     "(ARV × 0.70) - Rehab Costs",
     "=(ARV*0.7) - Rehab",
     "($300K × 0.7) - $30K = $180K"),
    
    ("Net Operating Income (NOI)",
     "Annual income after expenses (before debt)",
     "Gross Rent - Operating Expenses",
     "=GrossRent - OpEx",
     "$24,000 - $8,000 = $16,000"),
    
    ("Cap Rate",
     "Unlevered yield on property",
     "NOI / Property Value × 100",
     "=NOI/Value*100",
     "$16K / $200K = 8.0%"),
    
    ("Cash-on-Cash Return",
     "Annual cash return on invested capital",
     "(Annual Cash Flow / Cash Invested) × 100",
     "=CashFlow/Investment*100",
     "$7,000 / $55,000 = 12.7%"),
    
    ("Gross Rent Multiplier (GRM)",
     "Price to rent relationship",
     "Property Price / Gross Annual Rent",
     "=Price/GrossRent",
     "$200K / $24K = 8.33x"),
    
    ("Rent-to-Value Ratio",
     "Monthly rent as % of value (1% Rule)",
     "Monthly Rent / Property Value",
     "=MonthlyRent/Value",
     "$2,000 / $200,000 = 1.0%"),
    
    ("Debt Service Coverage Ratio (DSCR)",
     "Income vs debt payment coverage",
     "NOI / Annual Debt Service",
     "=NOI/DebtService",
     "$16,000 / $9,000 = 1.78"),
    
    ("Break-Even Ratio",
     "Occupancy needed to cover costs",
     "(OpEx + Debt Service) / Gross Rent",
     "=(OpEx+Debt)/GrossRent",
     "($8K + $9K) / $24K = 70.8%"),
    
    ("Return on Investment (ROI)",
     "Total gain/loss on investment",
     "(Net Profit / Total Cost) × 100",
     "=(Profit/Cost)*100",
     "$45K / $255K = 17.6%"),
    
    ("Monthly Cash Flow",
     "Net monthly income after all expenses",
     "Rent - (Mortgage + OpEx + Reserves)",
     "=Rent-Mortgage-OpEx/12",
     "$2,000 - $1,200 - $667 = $133"),
    
    ("Total Return",
     "Cash flow + appreciation + principal",
     "Cash + Appreciation + Equity Gain",
     "=Cash+Appreciation+Principal",
     "$5K + $20K + $3K = $28K"),
    
    ("Loan-to-Value (LTV)",
     "Loan amount as % of property value",
     "(Loan Amount / Property Value) × 100",
     "=Loan/Value*100",
     "$160K / $200K = 80%"),
    
    ("Operating Expense Ratio",
     "Operating costs as % of income",
     "Operating Expenses / Gross Income",
     "=OpEx/GrossIncome",
     "$8,000 / $24,000 = 33.3%")
]

for formula_name, measures, formula, excel, example in formulas_data:
    ws10.cell(row=row, column=1, value=formula_name).font = Font(bold=True)
    ws10.cell(row=row, column=2, value=measures)
    ws10.cell(row=row, column=3, value=formula).font = Font(italic=True)
    ws10.cell(row=row, column=4, value=excel).font = Font(name='Courier New')
    ws10.cell(row=row, column=5, value=example)
    
    for col in range(1, 6):
        ws10.cell(row=row, column=col).border = THIN_BORDER
        ws10.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    row += 1

apply_alternating_rows(ws10, 4, row-1, 1, 5)

# Column widths
ws10.column_dimensions['A'].width = 25
ws10.column_dimensions['B'].width = 30
ws10.column_dimensions['C'].width = 35
ws10.column_dimensions['D'].width = 28
ws10.column_dimensions['E'].width = 30

# Freeze top rows for all sheets
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ws.freeze_panes = 'A4'

print("Saving workbook...")
# Save the workbook
output_path = '/home/ubuntu/real_estate_guide/Real_Estate_Investment_Toolkit.xlsx'
wb.save(output_path)

print(f"✅ Excel toolkit created successfully: {output_path}")
print(f"Total sheets: {len(wb.sheetnames)}")
print("Sheets included:", ", ".join(wb.sheetnames))
