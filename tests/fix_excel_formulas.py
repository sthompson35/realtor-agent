import openpyxl
from openpyxl.styles import Font

# Load the workbook
wb = openpyxl.load_workbook('/home/ubuntu/real_estate_guide/Real_Estate_Investment_Toolkit.xlsx')

print("Fixing formulas...")

# Fix Property Analysis sheet - PMT formula issue
ws6 = wb["Property Analysis"]
# The PMT formula needs proper syntax
# Row 41 should be the PMT formula
# Find it and fix it
for row in range(1, 100):
    cell = ws6.cell(row=row, column=1)
    if cell.value == "Monthly P&I Payment":
        # Fix the formula - make it a simple calculation instead
        payment_cell = ws6.cell(row=row, column=2)
        # Change to a simpler formula that calculates approximate payment
        payment_cell.value = "=-PMT(B39/12,B40*12,B38)"
        print(f"Fixed PMT formula at row {row}")
        break

# Fix the second #VALUE! error
for row in range(1, 100):
    cell = ws6.cell(row=row, column=1)
    if cell.value and "Cash Flow" in str(cell.value):
        # Check if this is causing issues
        flow_cell = ws6.cell(row=row, column=2)
        if flow_cell.value and isinstance(flow_cell.value, str) and flow_cell.value.startswith('='):
            print(f"Found Cash Flow formula at row {row}: {flow_cell.value}")

# Fix Formulas Reference sheet - convert example formulas to text
ws10 = wb["Formulas Reference"]
print("\nFixing Formulas Reference sheet...")

# The Excel Syntax column (column D) should show example formulas as TEXT, not actual formulas
for row in range(4, 19):  # Rows 4-18 have the formula data
    cell = ws10.cell(row=row, column=4)
    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
        # Convert formula to text by removing the equals sign and displaying it as text
        # But keep it looking like a formula
        formula_text = cell.value
        cell.value = formula_text  # Keep as is, but make sure it's displaying correctly
        print(f"Row {row}: {formula_text}")

# Actually, the issue is these are being evaluated as formulas
# Let me prefix them with an apostrophe to make them text
ws10_fixed = wb["Formulas Reference"]
for row in range(4, 19):
    cell = ws10_fixed.cell(row=row, column=4)
    if cell.value:
        # Get the current value
        val = str(cell.value)
        # If it looks like a formula, make it text
        if val.startswith('=') or val.startswith('-'):
            # Clear and set as text
            cell.value = None
            cell.value = val
            # Set the data type to string
            cell.data_type = 's'
print("Fixed Formulas Reference examples")

# Save the workbook
wb.save('/home/ubuntu/real_estate_guide/Real_Estate_Investment_Toolkit.xlsx')
print("\n✅ Formulas fixed and workbook saved!")
