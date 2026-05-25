#!/usr/bin/env python3
"""
Realtor Agent Web Server — Phase 2
Flask application serving the dashboard and REST API.

© Shylow Thompson. LLC 2026 - All Rights Reserved
"""

from __future__ import annotations
import logging
import os
import threading
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, render_template, request, abort, session, g, redirect, url_for, send_file

# Ensure DB is initialised before first request
from realtor_agent.core.models import init_db
# Import database.py to register Activity/Appointment models with Base before create_all
import realtor_agent.core.database as _db_compat  # noqa: F401
from realtor_agent.web.db_queries import (
    get_dashboard_stats,
    get_deals,
    get_deal,
    get_pipeline,
    get_bot_status,
    get_recent_activity,
    get_system_status,
    get_analytics,
    get_closed_deals,
    BOT_NAMES,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# App factory
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
app = Flask(
    __name__,
    template_folder=str(BASE_DIR / "web" / "templates"),
    static_folder=str(BASE_DIR / "web" / "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-prod")

# Init DB tables on startup (idempotent)
with app.app_context():
    try:
        init_db()
        logger.info("Database initialised.")
    except Exception as exc:
        logger.error("DB init failed: %s", exc)


# ---------------------------------------------------------------------------
# APScheduler — auto-run pipeline every 6 hours
# ---------------------------------------------------------------------------

def _scheduled_pipeline_run():
    """Background task: full 8-bot pipeline every 6 hours."""
    logger.info("Scheduled pipeline run starting")
    try:
        from realtor_agent.core.orchestrator import Orchestrator
        orch = Orchestrator()
        results = orch.run()
        logger.info("Scheduled pipeline complete — %d bots ran", len(results))
    except Exception as exc:
        logger.error("Scheduled pipeline failed: %s", exc)


try:
    from realtor_agent.automation.scheduler import scheduled_task_manager, DefaultScheduledTasks
    scheduled_task_manager.start()
    # Pipeline every 6 hours
    scheduled_task_manager.add_interval_task(
        "pipeline_auto_run",
        _scheduled_pipeline_run,
        hours=6,
    )
    # Keep existing default tasks (daily cleanup, bot status check, lead scoring, market refresh)
    DefaultScheduledTasks.setup_default_tasks(scheduled_task_manager)
    logger.info("APScheduler started — pipeline will run every 6 hours")
except Exception as _sched_exc:
    logger.warning("APScheduler not started: %s", _sched_exc)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _json_error(msg: str, status: int = 400):
    return jsonify({"error": msg}), status


def _sanitize(obj):
    """Recursively replace inf/nan float values with None so JSON serialization doesn't break."""
    import math
    if isinstance(obj, float):
        return None if (math.isinf(obj) or math.isnan(obj)) else obj
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize(v) for v in obj]
    return obj


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

@app.route("/")
def dashboard():
    return render_template("dashboard.html")


@app.route("/deals")
@app.route("/deals/")
def deals_page():
    return render_template("deals.html")


@app.route("/deals/new")
def deals_new():
    return render_template("deals.html", show_new_deal=True)


@app.route("/deals/<int:deal_id>")
def deal_detail(deal_id: int):
    deal = get_deal(deal_id)
    if not deal:
        abort(404)
    return render_template("deals.html", deal=deal)


@app.route("/deals/pipeline")
def pipeline_page():
    return render_template("pipeline.html")


@app.route("/deals/archived")
def archived_page():
    return render_template("deals.html", show_archived=True)


@app.route("/bots")
def bots_page():
    return render_template("bots.html")


@app.route("/bots/run")
def bots_run_page():
    return render_template("bots.html", show_run_interface=True)


@app.route("/analytics")
def analytics_page():
    return render_template("analytics.html")


@app.route("/reports")
def reports_page():
    return render_template("reports.html")


@app.route("/reports/performance")
def reports_performance():
    return render_template("performance.html")


@app.route("/reports/market")
def reports_market():
    return render_template("market.html")


@app.route("/reports/financial")
def reports_financial():
    return render_template("financial.html")


@app.route("/reports/bot-activity")
def reports_bot_activity():
    return render_template("bot_activity.html")


@app.route("/login")
def login_page():
    return render_template("login.html")


@app.route("/register")
def register_page():
    return render_template("register.html")


@app.route("/logout")
def logout_page():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/contacts")
def contacts_page():
    return render_template("contacts.html")


@app.route("/settings")
def settings_page():
    return render_template("settings.html")


@app.route("/settings/bots")
def settings_bots():
    return render_template("settings.html", section="bots")


@app.route("/guide/land-flip")
def guide_land_flip():
    return render_template("land_flip_guide.html")


@app.route("/formulas")
def formulas_page():
    return render_template("formulas.html")


@app.route("/toolkit-dashboard")
def toolkit_dashboard():
    return render_template("toolkit_dashboard.html")


@app.route("/enhanced-deals")
def enhanced_deals():
    return render_template("enhanced_deals.html")


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@app.route("/health")
def health():
    return jsonify({"status": "healthy", "timestamp": datetime.utcnow().isoformat()})


# ---------------------------------------------------------------------------
# REST API — read endpoints
# ---------------------------------------------------------------------------

@app.route("/api/stats")
def api_stats():
    """Dashboard KPI cards."""
    return jsonify(get_dashboard_stats())


@app.route("/api/deals")
def api_deals():
    """Paginated deal list with optional filters."""
    status = request.args.get("status")
    search = request.args.get("search")
    limit = min(int(request.args.get("limit", 50)), 200)
    offset = int(request.args.get("offset", 0))
    return jsonify(get_deals(status=status, limit=limit, offset=offset, search=search))


@app.route("/api/deals/<int:deal_id>")
def api_deal(deal_id: int):
    """Single deal with full underwriting results."""
    deal = get_deal(deal_id)
    if not deal:
        return _json_error("Deal not found", 404)
    return jsonify(deal)


@app.route("/api/pipeline")
def api_pipeline():
    """Pipeline stage counts and deal cards."""
    return jsonify(get_pipeline())


@app.route("/api/bots/status")
def api_bots_status():
    """All bot last-run status."""
    return jsonify(get_bot_status())


@app.route("/api/recent-activity")
def api_recent_activity():
    """Sidebar activity feed."""
    limit = int(request.args.get("limit", 10))
    return jsonify(get_recent_activity(limit=limit))


@app.route("/api/system-status")
def api_system_status():
    """Sidebar system health."""
    return jsonify(get_system_status())


@app.route("/api/analytics")
def api_analytics():
    """Strategy performance and asset breakdown."""
    return jsonify(get_analytics())


@app.route("/api/closed-deals")
def api_closed_deals():
    """Closed deal archive."""
    limit = min(int(request.args.get("limit", 50)), 500)
    return jsonify(get_closed_deals(limit=limit))


# ---------------------------------------------------------------------------
# REST API — write / action endpoints
# ---------------------------------------------------------------------------

@app.route("/api/deals", methods=["POST"])
def api_create_deal():
    """
    Create a new deal intake record and immediately underwrite it.

    Body (JSON): all DealIntake fields. Required: address, contract_price, arv.
    """
    from realtor_agent.core.models import SessionLocal, DealIntakeModel
    from realtor_agent.bots.underwriter.underwriter import UnderwriterBot
    import uuid

    data = request.get_json(force=True) or {}
    if not data.get("address"):
        return _json_error("address is required")
    if not data.get("contract_price"):
        return _json_error("contract_price is required")

    # Assign a deal_id if not provided
    if not data.get("deal_id"):
        data["deal_id"] = f"D-{uuid.uuid4().hex[:8].upper()}"

    # Run underwriter bot (it handles DB persistence)
    bot = UnderwriterBot()
    result = bot.run({"web_scout": {"listings": [data]}})

    if result.get("results"):
        summary = result["results"][0]
        # Look up the new DB record id
        db = SessionLocal()
        try:
            record = db.query(DealIntakeModel).filter_by(deal_id=data["deal_id"]).first()
            db_id = record.id if record else None
        finally:
            db.close()

        return jsonify({
            "success": True,
            "deal_id": data["deal_id"],
            "db_id": db_id,
            "underwriting": summary,
        }), 201

    return _json_error("Failed to create deal"), 500


@app.route("/api/deals/<int:deal_id>", methods=["PATCH"])
def api_update_deal(deal_id: int):
    """Update deal status, priority, or notes."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel

    data = request.get_json(force=True) or {}
    allowed = {"status", "priority", "notes", "assigned_to"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return _json_error("No valid fields to update")

    db = SessionLocal()
    try:
        r = db.query(DealIntakeModel).filter(DealIntakeModel.id == deal_id).first()
        if not r:
            return _json_error("Deal not found", 404)
        for k, v in updates.items():
            setattr(r, k, v)
        r.updated_at = datetime.utcnow()
        db.commit()
        return jsonify({"success": True, "updated": updates})
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


@app.route("/api/deals/<int:deal_id>/underwrite", methods=["POST"])
def api_underwrite_deal(deal_id: int):
    """Re-run underwriting on an existing deal."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel, UnderwritingResultModel, RehabScopeModel
    from realtor_agent.bots.underwriter.underwriter import UnderwriterBot
    import dataclasses

    db = SessionLocal()
    try:
        r = db.query(DealIntakeModel).filter(DealIntakeModel.id == deal_id).first()
        if not r:
            return _json_error("Deal not found", 404)

        # Delete stale results so bot writes fresh ones
        db.query(UnderwritingResultModel).filter_by(deal_id_fk=deal_id).delete()
        db.query(RehabScopeModel).filter_by(deal_id_fk=deal_id).delete()
        db.commit()

        # Build raw dict the bot expects
        raw = {
            "deal_id": r.deal_id,
            "address": r.address,
            "city": r.city,
            "state": r.state,
            "contract_price": r.contract_price,
            "arv": r.arv,
            "rehab_cost": r.rehab_cost,
            "monthly_rent": r.monthly_rent,
            "taxes": r.taxes,
            "insurance": r.insurance,
            "hoa": r.hoa,
            "asset_type": r.asset_type,
            "acres": r.acres,
            "zoning": r.zoning,
            "buildable": r.buildable,
            "road_access": r.road_access,
            "flood_zone": r.flood_zone,
        }
    finally:
        db.close()

    bot = UnderwriterBot()
    result = bot.run({"web_scout": {"listings": [raw]}})
    return jsonify({"success": True, "underwriting": result.get("results", [{}])[0]})


@app.route("/api/bots/run", methods=["POST"])
def api_run_pipeline():
    """
    Kick off the full bot pipeline in a background thread.
    Returns immediately with a run_id; poll /api/bots/run/<run_id> for status.
    """
    from realtor_agent.core.orchestrator import Orchestrator
    from realtor_agent.core.models import SessionLocal, BotRun
    import uuid

    bot_names = request.get_json(force=True, silent=True) or {}
    bots = bot_names.get("bots")  # list or None (None = full sequence)

    run_id = uuid.uuid4().hex[:12]

    def _run():
        db = SessionLocal()
        run_record = BotRun(
            bot_name="pipeline",
            status="running",
            started_at=datetime.utcnow(),
        )
        db.add(run_record)
        db.commit()
        run_db_id = run_record.id
        db.close()

        try:
            orchestrator = Orchestrator(bot_names=bots)
            results = orchestrator.run()
            _RUN_RESULTS[run_id] = {
                "status": "success",
                "results": [r if isinstance(r, dict) else vars(r) for r in results],
                "completed_at": datetime.utcnow().isoformat(),
            }
            # Update DB record
            db2 = SessionLocal()
            rec = db2.query(BotRun).filter(BotRun.id == run_db_id).first()
            if rec:
                rec.status = "success"
                rec.completed_at = datetime.utcnow()
                rec.results = _RUN_RESULTS[run_id]
                db2.commit()
            db2.close()
        except Exception as exc:
            logger.error("Pipeline run failed: %s", exc)
            _RUN_RESULTS[run_id] = {
                "status": "failed",
                "error": str(exc),
                "completed_at": datetime.utcnow().isoformat(),
            }

    _RUN_RESULTS[run_id] = {"status": "running", "started_at": datetime.utcnow().isoformat()}
    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    return jsonify({"run_id": run_id, "status": "running"}), 202


_RUN_RESULTS: dict[str, dict] = {}


@app.route("/api/bots/run/<run_id>")
def api_run_status(run_id: str):
    """Poll pipeline run status."""
    result = _RUN_RESULTS.get(run_id)
    if not result:
        return _json_error("Run ID not found", 404)
    return jsonify(result)


@app.route("/api/bots/<bot_name>/run", methods=["POST"])
def api_run_bot(bot_name: str):
    """Run a single bot synchronously (for testing/manual triggers)."""
    if bot_name not in BOT_NAMES:
        return _json_error(f"Unknown bot: {bot_name}", 404)

    from realtor_agent.core.orchestrator import Orchestrator
    try:
        orchestrator = Orchestrator(bot_names=[bot_name])
        results = orchestrator.run()
        return jsonify({
            "bot": bot_name,
            "status": "success",
            "result": results[0] if results else {},
        })
    except Exception as exc:
        logger.error("Bot %s failed: %s", bot_name, exc)
        return _json_error(str(exc), 500)


# ---------------------------------------------------------------------------
# Underwriter quick-calc API  (no DB persistence — pure calculation)
# ---------------------------------------------------------------------------

@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    """
    Run all applicable strategy calculators on a deal payload.
    No DB writes — returns JSON results immediately.

    Body: same fields as DealIntake.
    """
    from realtor_agent.calculations.intake import DealIntake
    from realtor_agent.calculations import strategies
    from realtor_agent.calculations.rehab_engine import calc_rehab_scope, ScopeLevel
    from realtor_agent.bots.underwriter.underwriter import _run_strategy, _STRATEGY_MAP, _scope_level_from_rehab
    import dataclasses

    data = request.get_json(force=True) or {}

    try:
        deal = DealIntake(
            deal_id=data.get("deal_id", "CALC"),
            address=data.get("address", ""),
            city=data.get("city", ""),
            state=data.get("state", ""),
            asset_type=data.get("asset_type", "SFR"),
            contract_price=float(data.get("contract_price", 0)),
            arv=float(data.get("arv", 0)),
            rehab_cost=float(data.get("rehab_cost", 0)),
            monthly_rent=float(data.get("monthly_rent", 0)),
            taxes=float(data.get("taxes", 0)),
            insurance=float(data.get("insurance", 0)),
            hoa=float(data.get("hoa", 0)),
            hold_months=int(data.get("hold_months", 6)),
            rate=float(data.get("rate", 0.08)),
            down_pct=float(data.get("down_pct", 0.20)),
            acres=float(data.get("acres", 0)),
            zoning=data.get("zoning", ""),
            buildable=bool(data.get("buildable", True)),
            road_access=bool(data.get("road_access", True)),
            flood_zone=data.get("flood_zone", "X"),
            flip_profit_target_pct=float(data.get("flip_profit_target_pct", 0.20)),
        )
    except (TypeError, ValueError) as exc:
        return _json_error(f"Invalid field: {exc}")

    strategy_names = _STRATEGY_MAP.get(deal.asset_type, _STRATEGY_MAP["SFR"])
    results = {}
    for name in strategy_names:
        r = _run_strategy(name, deal)
        if r:
            results[name] = dataclasses.asdict(r)

    scope = calc_rehab_scope(
        scope_level=_scope_level_from_rehab(deal.rehab_cost, deal.arv),
        sqft=float(data.get("sqft", 1200)),
    )

    return jsonify(_sanitize({
        "deal": {
            "address": deal.address,
            "asset_type": deal.asset_type,
            "contract_price": deal.contract_price,
            "arv": deal.arv,
            "rehab_cost": deal.rehab_cost,
            "noi": deal.noi,
        },
        "strategies": results,
        "rehab_estimate": {
            "scope_level": scope.scope_level,
            "subtotal": scope.subtotal,
            "contingency": scope.contingency_amount,
            "total": scope.total,
            "line_item_count": len(scope.line_items),
        },
    }))


# ---------------------------------------------------------------------------
# Closed deals archive — POST (import from Excel export)
# ---------------------------------------------------------------------------

@app.route("/api/closed-deals", methods=["POST"])
def api_import_closed_deal():
    """Import a single closed deal record (mirrors archive companion fields)."""
    from realtor_agent.core.models import SessionLocal, ClosedDealModel
    from realtor_agent.calculations.portfolio import calc_deal_metrics, ClosedDeal
    from datetime import date

    data = request.get_json(force=True) or {}
    if not data.get("deal_id"):
        return _json_error("deal_id is required")

    def _parse_date(v):
        if not v:
            return None
        try:
            return date.fromisoformat(v[:10])
        except Exception:
            return None

    cd = ClosedDeal(
        deal_id=data["deal_id"],
        close_date=_parse_date(data.get("close_date")),
        exit_date=_parse_date(data.get("exit_date")),
        market=data.get("market", ""),
        strategy=data.get("strategy", ""),
        entity=data.get("entity", ""),
        status=data.get("status", "Closed"),
        proj_purchase=float(data.get("proj_purchase", 0)),
        proj_rehab=float(data.get("proj_rehab", 0)),
        proj_exit=float(data.get("proj_exit", 0)),
        proj_rent=float(data.get("proj_rent", 0)),
        actual_purchase=float(data.get("actual_purchase", 0)),
        actual_rehab=float(data.get("actual_rehab", 0)),
        actual_exit=float(data.get("actual_exit", 0)),
        actual_rent=float(data.get("actual_rent", 0)),
        debt_proceeds=float(data.get("debt_proceeds", 0)),
        equity_in=float(data.get("equity_in", 0)),
        hold_months=int(data.get("hold_months", 0)),
    )
    metrics = calc_deal_metrics(cd)

    db = SessionLocal()
    try:
        existing = db.query(ClosedDealModel).filter_by(deal_id=cd.deal_id).first()
        if existing:
            return _json_error(f"Deal {cd.deal_id} already exists — use PATCH to update", 409)

        row = ClosedDealModel(
            deal_id=cd.deal_id,
            close_date=cd.close_date,
            exit_date=cd.exit_date,
            market=cd.market,
            strategy=cd.strategy,
            entity=cd.entity,
            status=cd.status,
            proj_purchase=cd.proj_purchase,
            proj_rehab=cd.proj_rehab,
            proj_exit=cd.proj_exit,
            proj_rent=cd.proj_rent,
            actual_purchase=cd.actual_purchase,
            actual_rehab=cd.actual_rehab,
            actual_exit=cd.actual_exit,
            actual_rent=cd.actual_rent,
            debt_proceeds=cd.debt_proceeds,
            equity_in=cd.equity_in,
            hold_months=cd.hold_months,
            gross_profit=metrics.gross_profit,
            net_profit=metrics.net_profit,
            roi=metrics.roi,
            coc=metrics.coc,
            dscr=metrics.dscr,
            purchase_variance_pct=metrics.proj_vs_actual_purchase,
            rehab_variance_pct=metrics.proj_vs_actual_rehab,
            exit_variance_pct=metrics.proj_vs_actual_exit,
        )
        db.add(row)
        db.commit()
        return jsonify({
            "success": True,
            "deal_id": cd.deal_id,
            "gross_profit": metrics.gross_profit,
            "net_profit": metrics.net_profit,
            "roi": round(metrics.roi * 100, 1),
        }), 201
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Lead generation / auto-find endpoints (used by deals.html actions)
# ---------------------------------------------------------------------------

@app.route("/api/leads/generate", methods=["POST"])
@app.route("/api/leads/find-new", methods=["POST"])
@app.route("/api/leads/generate-ai", methods=["POST"])
def api_generate_leads():
    """
    Trigger the web_scout + data_clean + underwriter mini-pipeline to
    generate and score a batch of new leads.  Returns immediately with a
    run_id for polling (same pattern as /api/bots/run).
    """
    import uuid
    from realtor_agent.core.orchestrator import Orchestrator
    from realtor_agent.core.models import SessionLocal, BotRun

    data      = request.get_json(force=True, silent=True) or {}
    num_leads = int(data.get("num_leads") or data.get("max_leads") or data.get("count") or 10)
    num_leads = max(1, min(num_leads, 50))
    run_id    = uuid.uuid4().hex[:12]

    def _run():
        db  = SessionLocal()
        rec = BotRun(bot_name="lead_gen", status="running", started_at=datetime.utcnow())
        db.add(rec); db.commit(); run_db_id = rec.id; db.close()

        try:
            orch = Orchestrator(bot_names=["web_scout", "data_clean", "underwriter"])
            orch.context["num_leads"] = num_leads
            results = orch.run()
            uw      = results[2] if len(results) > 2 else {}
            payload = {
                "status":        "completed",
                "leads_generated": uw.get("total", 0),
                "approved":      uw.get("approved_count", 0),
                "run_id":        run_id,
            }
            _RUN_RESULTS[run_id] = payload

            db2  = SessionLocal()
            rec2 = db2.query(BotRun).filter(BotRun.id == run_db_id).first()
            if rec2:
                rec2.status = "success"; rec2.completed_at = datetime.utcnow()
                rec2.results = payload; db2.commit()
            db2.close()
        except Exception as exc:
            logger.error("Lead gen run failed: %s", exc)
            _RUN_RESULTS[run_id] = {"status": "error", "error": str(exc)}

    _RUN_RESULTS[run_id] = {"status": "running"}
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({
        "success": True,
        "run_id":  run_id,
        "status":  "running",
        "message": f"Generating {num_leads} leads in background — poll /api/bots/run/{run_id}",
    }), 202


@app.route("/api/deals/auto-find", methods=["POST"])
def api_auto_find_deals():
    """
    Run the full 8-bot pipeline to discover, underwrite, and queue outreach
    for profitable deals matching the supplied criteria.
    Returns run_id for polling.
    """
    import uuid
    from realtor_agent.core.orchestrator import Orchestrator
    from realtor_agent.core.models import SessionLocal, BotRun

    data      = request.get_json(force=True, silent=True) or {}
    num_leads = int(data.get("max_deals") or data.get("num_leads") or 10)
    num_leads = max(1, min(num_leads, 50))
    run_id    = uuid.uuid4().hex[:12]

    def _run():
        db  = SessionLocal()
        rec = BotRun(bot_name="auto_find", status="running", started_at=datetime.utcnow())
        db.add(rec); db.commit(); run_db_id = rec.id; db.close()

        try:
            orch = Orchestrator()
            orch.context["num_leads"] = num_leads
            results = orch.run()
            uw      = next((r for r in results if r.get("approved_count") is not None), {})
            neg     = next((r for r in results if r.get("advanced") is not None), {})
            payload = {
                "status":          "completed",
                "leads_generated": uw.get("total", 0),
                "profitable_deals": uw.get("approved_count", 0),
                "deals_advanced":  neg.get("advanced", 0),
                "run_id":          run_id,
            }
            _RUN_RESULTS[run_id] = payload

            db2  = SessionLocal()
            rec2 = db2.query(BotRun).filter(BotRun.id == run_db_id).first()
            if rec2:
                rec2.status = "success"; rec2.completed_at = datetime.utcnow()
                rec2.results = payload; db2.commit()
            db2.close()
        except Exception as exc:
            logger.error("Auto-find run failed: %s", exc)
            _RUN_RESULTS[run_id] = {"status": "error", "error": str(exc)}

    _RUN_RESULTS[run_id] = {"status": "running"}
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({
        "success": True,
        "run_id":  run_id,
        "status":  "running",
        "message": f"Auto-finding deals in background — poll /api/bots/run/{run_id}",
    }), 202


@app.route("/api/market/scrape-data", methods=["POST"])
def api_scrape_market():
    """Alias: trigger lead generation (web_scout) scoped to passed markets."""
    return api_generate_leads()


@app.route("/api/schedule", methods=["GET"])
def api_schedule_list():
    """Return current scheduled tasks."""
    try:
        from realtor_agent.automation.scheduler import scheduled_task_manager
        tasks = scheduled_task_manager.get_tasks()
        serialised = {}
        for tid, t in tasks.items():
            serialised[tid] = {
                "type":     t.get("type"),
                "function": t.get("function"),
                "trigger":  str(t.get("trigger", "")),
                "next_run": t.get("next_run").isoformat() if t.get("next_run") else None,
            }
        return jsonify(serialised)
    except Exception as exc:
        return jsonify({"error": str(exc), "tasks": {}})


# ---------------------------------------------------------------------------
# Phase 4 — Auth  (POST /api/auth/*)
# ---------------------------------------------------------------------------

def _get_current_user():
    """Return decoded token payload from Bearer header or session, or None."""
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
        from realtor_agent.core.auth import auth_service
        return auth_service.verify_token(token)
    return session.get("user")


@app.route("/api/auth/register", methods=["POST"])
def api_auth_register():
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip()
    email    = (data.get("email") or "").strip()
    password = data.get("password") or ""
    if not username or not email or not password:
        return _json_error("username, email, and password are required")
    from realtor_agent.core.auth import auth_service
    result = auth_service.create_user(username, email, password, role=data.get("role", "agent"))
    if not result:
        return _json_error("Username or email already exists", 409)
    return jsonify({"success": True, "user": result}), 201


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data     = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or not password:
        return _json_error("username and password are required")
    from realtor_agent.core.auth import auth_service
    result = auth_service.authenticate_user(username, password)
    if not result:
        return _json_error("Invalid credentials or account locked", 401)
    # Store minimal user info in server-side session too
    session["user"] = {
        "user_id":  result["user"]["id"],
        "username": result["user"]["username"],
        "role":     result["user"]["role"],
    }
    return jsonify({"success": True, "token": result["token"], "user": result["user"]})


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/auth/me", methods=["GET"])
def api_auth_me():
    user = _get_current_user()
    if not user:
        return _json_error("Not authenticated", 401)
    return jsonify(user)


# ---------------------------------------------------------------------------
# Phase 4 — Contacts  (GET/POST/PATCH/DELETE /api/contacts/*)
# ---------------------------------------------------------------------------

@app.route("/api/contacts", methods=["GET"])
def api_contacts_list():
    from realtor_agent.core.models import SessionLocal, Contact
    db = SessionLocal()
    try:
        q = db.query(Contact)
        if request.args.get("type"):
            q = q.filter(Contact.contact_type == request.args["type"])
        if request.args.get("deal_id"):
            q = q.filter(Contact.deal_id_fk == int(request.args["deal_id"]))
        search = request.args.get("q", "").strip()
        if search:
            like = f"%{search}%"
            q = q.filter(
                Contact.name.ilike(like) |
                Contact.email.ilike(like) |
                Contact.phone.ilike(like)
            )
        if request.args.get("dnc") == "false":
            q = q.filter(Contact.dnc == False)
        limit  = min(int(request.args.get("limit", 100)), 500)
        offset = int(request.args.get("offset", 0))
        total  = q.count()
        rows   = q.order_by(Contact.created_at.desc()).offset(offset).limit(limit).all()
        return jsonify({
            "contacts": [_contact_to_dict(c) for c in rows],
            "total": total,
            "limit": limit,
            "offset": offset,
        })
    finally:
        db.close()


@app.route("/api/contacts", methods=["POST"])
def api_contacts_create():
    from realtor_agent.core.models import SessionLocal, Contact
    data = request.get_json(force=True, silent=True) or {}
    if not data.get("name"):
        return _json_error("name is required")
    db = SessionLocal()
    try:
        c = Contact(
            deal_id_fk    = data.get("deal_id"),
            contact_type  = data.get("contact_type", "owner"),
            name          = data["name"].strip(),
            email         = data.get("email", "").strip() or None,
            phone         = data.get("phone", "").strip() or None,
            mailing_address = data.get("mailing_address"),
            ownership_type  = data.get("ownership_type"),
            dnc           = bool(data.get("dnc", False)),
            consent_given = bool(data.get("consent_given", False)),
            skip_traced   = bool(data.get("skip_traced", False)),
            notes         = data.get("notes"),
        )
        db.add(c)
        db.commit()
        db.refresh(c)
        _push_notification(db, "New contact added", f"{c.name} added to contacts", "success", "contact", f"/contacts")
        return jsonify({"success": True, "contact": _contact_to_dict(c)}), 201
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc))
    finally:
        db.close()


@app.route("/api/contacts/<int:contact_id>", methods=["GET"])
def api_contacts_get(contact_id):
    from realtor_agent.core.models import SessionLocal, Contact
    db = SessionLocal()
    try:
        c = db.query(Contact).get(contact_id)
        if not c:
            return _json_error("Contact not found", 404)
        return jsonify(_contact_to_dict(c))
    finally:
        db.close()


@app.route("/api/contacts/<int:contact_id>", methods=["PATCH"])
def api_contacts_update(contact_id):
    from realtor_agent.core.models import SessionLocal, Contact
    data = request.get_json(force=True, silent=True) or {}
    db   = SessionLocal()
    try:
        c = db.query(Contact).get(contact_id)
        if not c:
            return _json_error("Contact not found", 404)
        for field in ("name", "email", "phone", "mailing_address", "ownership_type",
                      "dnc", "consent_given", "skip_traced", "notes", "contact_type"):
            if field in data:
                setattr(c, field, data[field])
        db.commit()
        return jsonify({"success": True, "contact": _contact_to_dict(c)})
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc))
    finally:
        db.close()


@app.route("/api/contacts/<int:contact_id>", methods=["DELETE"])
def api_contacts_delete(contact_id):
    from realtor_agent.core.models import SessionLocal, Contact
    db = SessionLocal()
    try:
        c = db.query(Contact).get(contact_id)
        if not c:
            return _json_error("Contact not found", 404)
        db.delete(c)
        db.commit()
        return jsonify({"success": True})
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc))
    finally:
        db.close()


def _contact_to_dict(c):
    return {
        "id":             c.id,
        "deal_id":        c.deal_id_fk,
        "contact_type":   c.contact_type,
        "name":           c.name,
        "email":          c.email,
        "phone":          c.phone,
        "mailing_address": c.mailing_address,
        "ownership_type": c.ownership_type,
        "dnc":            c.dnc,
        "consent_given":  c.consent_given,
        "skip_traced":    c.skip_traced,
        "notes":          c.notes,
        "created_at":     c.created_at.isoformat() if c.created_at else None,
        "updated_at":     c.updated_at.isoformat() if c.updated_at else None,
    }


# ---------------------------------------------------------------------------
# Phase 4 — Deal Export  (GET /api/deals/export)
# ---------------------------------------------------------------------------

@app.route("/api/deals/export", methods=["GET"])
def api_deals_export():
    import csv, io
    from realtor_agent.core.models import SessionLocal, DealIntakeModel

    fmt     = request.args.get("format", "csv").lower()
    status  = request.args.get("status")
    db      = SessionLocal()
    try:
        q = db.query(DealIntakeModel)
        if status:
            q = q.filter(DealIntakeModel.status == status)
        deals = q.order_by(DealIntakeModel.created_at.desc()).all()

        COLS = [
            "deal_id", "address", "city", "state", "zip_code", "asset_type",
            "contract_price", "arv", "rehab_cost", "monthly_rent",
            "taxes", "insurance", "hoa", "status", "priority", "source",
            "assigned_to", "score", "deal_score", "acres", "zoning",
            "flood_zone", "buildable", "notes", "created_at",
        ]

        if fmt == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                return _json_error("openpyxl not installed — use format=csv", 501)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deals"

            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx, col in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, d in enumerate(deals, 2):
                for col_idx, col in enumerate(COLS, 1):
                    val = getattr(d, col, None)
                    if hasattr(val, "isoformat"):
                        val = val.isoformat()
                    ws.cell(row=row_idx, column=col_idx, value=val)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="deals_export.xlsx",
            )

        # Default: CSV
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([c.replace("_", " ").title() for c in COLS])
        for d in deals:
            writer.writerow([
                str(getattr(d, col, "") or "") for col in COLS
            ])
        buf.seek(0)
        return send_file(
            io.BytesIO(buf.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="deals_export.csv",
        )
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Phase 4 — Notifications  (GET/POST/PATCH /api/notifications/*)
# ---------------------------------------------------------------------------

def _push_notification(db, title: str, message: str, level: str = "info",
                        category: str = "system", link: str = None):
    """Insert a notification row. Caller owns the db session."""
    try:
        from realtor_agent.core.models import Notification
        n = Notification(title=title, message=message, level=level,
                         category=category, link=link)
        db.add(n)
        # commit is caller's responsibility
    except Exception as exc:
        logger.warning("Could not create notification: %s", exc)


@app.route("/api/notifications", methods=["GET"])
def api_notifications_list():
    from realtor_agent.core.models import SessionLocal, Notification
    db = SessionLocal()
    try:
        q = db.query(Notification)
        if request.args.get("unread") == "true":
            q = q.filter(Notification.read == False)
        limit = min(int(request.args.get("limit", 50)), 200)
        rows  = q.order_by(Notification.created_at.desc()).limit(limit).all()
        return jsonify([_notif_to_dict(n) for n in rows])
    finally:
        db.close()


@app.route("/api/notifications/count", methods=["GET"])
def api_notifications_count():
    from realtor_agent.core.models import SessionLocal, Notification
    db = SessionLocal()
    try:
        count = db.query(Notification).filter(Notification.read == False).count()
        return jsonify({"unread": count})
    finally:
        db.close()


@app.route("/api/notifications", methods=["POST"])
def api_notifications_create():
    from realtor_agent.core.models import SessionLocal, Notification
    data = request.get_json(force=True, silent=True) or {}
    if not data.get("title") or not data.get("message"):
        return _json_error("title and message are required")
    db = SessionLocal()
    try:
        n = Notification(
            title    = data["title"],
            message  = data["message"],
            level    = data.get("level", "info"),
            category = data.get("category", "system"),
            link     = data.get("link"),
        )
        db.add(n)
        db.commit()
        db.refresh(n)
        return jsonify({"success": True, "notification": _notif_to_dict(n)}), 201
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc))
    finally:
        db.close()


@app.route("/api/notifications/<int:notif_id>/read", methods=["PATCH"])
def api_notifications_mark_read(notif_id):
    from realtor_agent.core.models import SessionLocal, Notification
    db = SessionLocal()
    try:
        n = db.query(Notification).get(notif_id)
        if not n:
            return _json_error("Notification not found", 404)
        n.read = True
        db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@app.route("/api/notifications/read-all", methods=["PATCH"])
def api_notifications_read_all():
    from realtor_agent.core.models import SessionLocal, Notification
    db = SessionLocal()
    try:
        db.query(Notification).filter(Notification.read == False).update({"read": True})
        db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


def _notif_to_dict(n):
    return {
        "id":         n.id,
        "title":      n.title,
        "message":    n.message,
        "level":      n.level,
        "category":   n.category,
        "link":       n.link,
        "read":       n.read,
        "created_at": n.created_at.isoformat() if n.created_at else None,
    }


# ---------------------------------------------------------------------------
# Phase 4 — User Settings  (GET/PATCH /api/settings)
# ---------------------------------------------------------------------------

_DEFAULT_SETTINGS = {
    # Underwriting defaults
    "flip_profit_target_pct":  0.20,
    "wholesale_min_fee":       5000.0,
    "investor_margin_pct":     0.15,
    "default_hold_months":     6,
    "default_vacancy_pct":     0.08,
    "default_pm_pct":          0.08,
    "default_repairs_pct":     0.05,
    "default_capex_pct":       0.05,
    # Pipeline
    "pipeline_auto_run":       True,
    "pipeline_interval_hours": 6,
    "leads_per_run":           10,
    # Notifications
    "notify_new_deal":         True,
    "notify_bot_complete":     True,
    "notify_deal_status":      True,
    "notify_email":            False,
    "notify_sms":              False,
    # UI
    "default_asset_type":      "SFR",
    "target_markets":          [],
    "dark_mode":               False,
}


@app.route("/api/settings", methods=["GET"])
def api_settings_get():
    from realtor_agent.core.models import SessionLocal, User
    user_ctx = _get_current_user()
    if user_ctx:
        db = SessionLocal()
        try:
            u = db.query(User).get(user_ctx.get("user_id"))
            if u and u.settings_json:
                merged = {**_DEFAULT_SETTINGS, **u.settings_json}
                return jsonify(merged)
        finally:
            db.close()
    return jsonify(_DEFAULT_SETTINGS)


@app.route("/api/settings", methods=["POST", "PATCH"])
def api_settings_update():
    from realtor_agent.core.models import SessionLocal, User
    data = request.get_json(force=True, silent=True) or {}
    # Filter to known keys only to prevent injection of arbitrary data
    allowed = set(_DEFAULT_SETTINGS.keys())
    clean   = {k: v for k, v in data.items() if k in allowed}
    if not clean:
        return _json_error("No valid settings keys provided")

    user_ctx = _get_current_user()
    if user_ctx:
        db = SessionLocal()
        try:
            u = db.query(User).get(user_ctx.get("user_id"))
            if u:
                existing = u.settings_json or {}
                existing.update(clean)
                u.settings_json = existing
                db.commit()
                return jsonify({"success": True, "settings": {**_DEFAULT_SETTINGS, **existing}})
        finally:
            db.close()

    # No auth — return the merged result without persisting
    return jsonify({"success": True, "settings": {**_DEFAULT_SETTINGS, **clean},
                    "warning": "Not authenticated — settings not persisted"})


# ---------------------------------------------------------------------------
# Phase 4 — Reports  (GET /api/reports/*)
# ---------------------------------------------------------------------------

@app.route("/api/reports/performance", methods=["GET"])
def api_reports_performance():
    from realtor_agent.core.models import SessionLocal, DealIntakeModel, BotRun
    from sqlalchemy import func
    db = SessionLocal()
    try:
        # Stage breakdown
        stage_counts = (
            db.query(DealIntakeModel.status, func.count(DealIntakeModel.id))
            .group_by(DealIntakeModel.status).all()
        )
        stage_map = {s: c for s, c in stage_counts}
        total     = sum(stage_map.values()) or 1

        # Conversion: leads → qualified → offer_sent → under_contract → closed
        FUNNEL = ["lead", "qualified", "offer_sent", "under_contract", "closed"]
        funnel  = [{"stage": s, "count": stage_map.get(s, 0)} for s in FUNNEL]

        lead_count   = stage_map.get("lead", 0)
        closed_count = stage_map.get("closed", 0)
        conv_rate    = round(closed_count / max(total, 1) * 100, 1)

        # Avg deal score and deal_score
        avg_score = db.query(func.avg(DealIntakeModel.deal_score)).scalar() or 0

        # Bot run stats
        bot_runs_total   = db.query(BotRun).count()
        bot_runs_success = db.query(BotRun).filter(BotRun.status == "success").count()
        bot_success_rate = round(bot_runs_success / max(bot_runs_total, 1) * 100, 1)

        # Recent activity: deals created last 30 days
        from datetime import timedelta
        cutoff = datetime.utcnow() - timedelta(days=30)
        recent_deals = db.query(DealIntakeModel).filter(
            DealIntakeModel.created_at >= cutoff
        ).count()

        return jsonify({
            "funnel":            funnel,
            "stage_breakdown":   stage_map,
            "total_deals":       total,
            "closed_deals":      closed_count,
            "conversion_rate":   conv_rate,
            "avg_deal_score":    round(float(avg_score), 2),
            "bot_runs_total":    bot_runs_total,
            "bot_success_rate":  bot_success_rate,
            "deals_last_30_days": recent_deals,
            "generated_at":      datetime.utcnow().isoformat(),
        })
    finally:
        db.close()


@app.route("/api/reports/financial", methods=["GET"])
def api_reports_financial():
    from realtor_agent.core.models import SessionLocal, DealIntakeModel, ClosedDealModel, UnderwritingResultModel
    from sqlalchemy import func
    db = SessionLocal()
    try:
        # Pipeline value (active deals, contract price sum)
        pipeline_val = db.query(func.sum(DealIntakeModel.contract_price)).filter(
            DealIntakeModel.status.notin_(["dead", "closed"])
        ).scalar() or 0

        # Closed deal totals
        total_invested = db.query(
            func.sum(ClosedDealModel.actual_purchase + ClosedDealModel.actual_rehab)
        ).scalar() or 0
        total_exit     = db.query(func.sum(ClosedDealModel.actual_exit)).scalar() or 0
        total_profit   = db.query(func.sum(ClosedDealModel.net_profit)).scalar() or 0
        avg_roi        = db.query(func.avg(ClosedDealModel.roi)).scalar() or 0
        total_closed   = db.query(ClosedDealModel).count()

        # Best strategy (most viable underwriting rows)
        top_strategies = (
            db.query(
                UnderwritingResultModel.strategy,
                func.count(UnderwritingResultModel.id).label("cnt"),
                func.avg(UnderwritingResultModel.roi).label("avg_roi"),
            )
            .filter(UnderwritingResultModel.viable == True)
            .group_by(UnderwritingResultModel.strategy)
            .order_by(func.count(UnderwritingResultModel.id).desc())
            .limit(5).all()
        )

        # Deals by asset type
        asset_breakdown = (
            db.query(DealIntakeModel.asset_type, func.count(DealIntakeModel.id))
            .group_by(DealIntakeModel.asset_type).all()
        )

        return jsonify({
            "pipeline_value":      round(float(pipeline_val), 2),
            "total_invested":      round(float(total_invested), 2),
            "total_exit_proceeds": round(float(total_exit), 2),
            "total_net_profit":    round(float(total_profit), 2),
            "avg_roi_pct":         round(float(avg_roi or 0) * 100, 2),
            "closed_deals_count":  total_closed,
            "top_strategies":      [
                {"strategy": s, "count": c, "avg_roi_pct": round(float(r or 0) * 100, 2)}
                for s, c, r in top_strategies
            ],
            "asset_breakdown":     {a: c for a, c in asset_breakdown},
            "generated_at":        datetime.utcnow().isoformat(),
        })
    finally:
        db.close()


@app.route("/api/reports/summary", methods=["GET"])
def api_reports_summary():
    """Executive summary combining performance + financial + recent bot activity."""
    perf_resp    = api_reports_performance().get_json()
    fin_resp     = api_reports_financial().get_json()
    stats_resp   = get_dashboard_stats()
    return jsonify({
        "performance": perf_resp,
        "financial":   fin_resp,
        "stats":       stats_resp,
        "generated_at": datetime.utcnow().isoformat(),
    })


# ---------------------------------------------------------------------------
# Phase 5 + Training Workbook — Calculation API  (POST /api/calculate/*)
# ---------------------------------------------------------------------------

@app.route("/api/calculate/sensitivity", methods=["POST"])
def api_calc_sensitivity():
    """7×7 grid: purchase ±15% vs ARV ±15% (5-pct steps). BUY/HOLD/PASS per cell."""
    from realtor_agent.calculations.advanced import calc_sensitivity_grid
    d = request.get_json(force=True) or {}
    try:
        result = calc_sensitivity_grid(
            base_purchase=float(d["contract_price"]),
            arv=float(d["arv"]),
            rehab=float(d.get("rehab_cost", 0)),
            hold_months=int(d.get("hold_months", 6)),
            sell_cost_pct=float(d.get("sell_cost_pct", 0.08)),
            buy_close_pct=float(d.get("buy_close_pct", 0.03)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/stress-test", methods=["POST"])
def api_calc_stress_test():
    """Best / Base / Worst scenarios with break-even months."""
    from realtor_agent.calculations.advanced import calc_stress_scenarios
    d = request.get_json(force=True) or {}
    try:
        result = calc_stress_scenarios(
            purchase=float(d["contract_price"]),
            arv=float(d["arv"]),
            rehab=float(d.get("rehab_cost", 0)),
            monthly_rent=float(d.get("monthly_rent", 0)),
            hold_months=int(d.get("hold_months", 6)),
            sell_cost_pct=float(d.get("sell_cost_pct", 0.08)),
            buy_close_pct=float(d.get("buy_close_pct", 0.03)),
            hold_cost_monthly=float(d.get("hold_cost_monthly", 0)),
            vacancy_pct=float(d.get("vacancy_pct", 0.08)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/jv", methods=["POST"])
def api_calc_jv():
    """JV waterfall: management fee → preferred returns → profit split."""
    from realtor_agent.calculations.advanced import calc_jv_waterfall
    d = request.get_json(force=True) or {}
    try:
        result = calc_jv_waterfall(
            total_profit=float(d["total_profit"]),
            partners=d["partners"],
            total_capital=float(d["total_capital"]),
            management_fee_pct=float(d.get("management_fee_pct", 0.02)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/exit-strategy", methods=["POST"])
def api_calc_exit_strategy():
    """Flip, Wholesale, Hold, BRRRR side-by-side with RECOMMENDED label."""
    from realtor_agent.calculations.advanced import calc_exit_comparison
    d = request.get_json(force=True) or {}
    try:
        result = calc_exit_comparison(
            purchase=float(d["contract_price"]),
            arv=float(d["arv"]),
            rehab=float(d.get("rehab_cost", 0)),
            monthly_rent=float(d.get("monthly_rent", 0)),
            hold_months=int(d.get("hold_months", 6)),
            sell_cost_pct=float(d.get("sell_cost_pct", 0.08)),
            buy_close_pct=float(d.get("buy_close_pct", 0.03)),
            rate=float(d.get("rate", 0.08)),
            down_pct=float(d.get("down_pct", 0.20)),
            term_years=int(d.get("term_years", 30)),
            vacancy_pct=float(d.get("vacancy_pct", 0.08)),
            pm_pct=float(d.get("pm_pct", 0.08)),
            repairs_pct=float(d.get("repairs_pct", 0.05)),
            capex_pct=float(d.get("capex_pct", 0.05)),
            taxes=float(d.get("taxes", 0)),
            insurance=float(d.get("insurance", 0)),
            hoa=float(d.get("hoa", 0)),
            refi_ltv=float(d.get("refi_ltv", 0.75)),
            wholesale_fee=float(d.get("wholesale_min_fee", 5000)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/wholesale", methods=["POST"])
def api_calc_wholesale():
    """70% rule, 65% rule, buyer-profit MAO. Returns GO/CAUTION/NO-GO."""
    from realtor_agent.calculations.advanced import calc_wholesale_mao
    d = request.get_json(force=True) or {}
    try:
        result = calc_wholesale_mao(
            arv=float(d["arv"]),
            rehab=float(d.get("rehab_cost", 0)),
            buyer_target_profit_pct=float(d.get("buyer_target_profit_pct", 0.20)),
            assignment_fee=float(d.get("assignment_fee", 5000)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/taxes", methods=["POST"])
def api_calc_taxes():
    """Federal 24%, MO 5%, SE 15.3% (flip). Depreciation shelter for holds."""
    from realtor_agent.calculations.advanced import calc_taxes
    d = request.get_json(force=True) or {}
    try:
        result = calc_taxes(
            net_profit=float(d["net_profit"]),
            strategy=d.get("strategy", "flip"),
            purchase_price=float(d.get("purchase_price", d.get("contract_price", 0))),
            hold_years=float(d.get("hold_years", 1.0)),
            land_value_pct=float(d.get("land_value_pct", 0.20)),
            federal_rate=float(d.get("federal_rate", 0.24)),
            state_rate=float(d.get("state_rate", 0.05)),
            se_tax_rate=float(d.get("se_tax_rate", 0.153)),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/draw-schedule", methods=["POST"])
def api_calc_draw_schedule():
    """6-phase construction draw schedule with running totals + contingency."""
    from realtor_agent.calculations.advanced import calc_draw_schedule
    d = request.get_json(force=True) or {}
    try:
        result = calc_draw_schedule(
            total_rehab=float(d["rehab_cost"]),
            contingency_pct=float(d.get("contingency_pct", 0.10)),
            phases=d.get("phases"),
        )
        return jsonify(result)
    except (KeyError, ValueError, TypeError) as exc:
        return _json_error(f"Missing/invalid field: {exc}")


@app.route("/api/calculate/comps", methods=["POST"])
def api_calc_comps():
    """5-comp analysis with sqft/bed/bath/age/condition adjustments, ±15% cap."""
    from realtor_agent.calculations.advanced import calc_comp_analysis
    d = request.get_json(force=True) or {}
    subject = d.get("subject") or {}
    comps   = d.get("comps") or []
    if not comps:
        return _json_error("comps list is required")
    result = calc_comp_analysis(
        subject=subject,
        comps=comps,
        price_per_sqft_adj=float(d.get("price_per_sqft_adj", 20.0)),
        bed_adj=float(d.get("bed_adj", 3000.0)),
        bath_adj=float(d.get("bath_adj", 2000.0)),
        age_adj=float(d.get("age_adj", 500.0)),
        condition_adj=float(d.get("condition_adj", 5000.0)),
    )
    return jsonify(result)


@app.route("/api/calculate/compare", methods=["POST"])
def api_calc_compare():
    """Composite ranking matrix: ROI 35%, Cash Flow 25%, Equity 20%, Cap Rate 20%."""
    from realtor_agent.calculations.advanced import calc_property_comparison
    d = request.get_json(force=True) or {}
    props = d.get("properties") or []
    if not props:
        return _json_error("properties list is required")
    return jsonify(calc_property_comparison(props))


@app.route("/api/pipeline/metrics", methods=["GET"])
def api_pipeline_metrics():
    """Probability-weighted pipeline value from DB stage counts."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel
    from realtor_agent.calculations.advanced import calc_pipeline_metrics
    from sqlalchemy import func

    db = SessionLocal()
    try:
        rows = (
            db.query(DealIntakeModel.status,
                     func.count(DealIntakeModel.id).label("cnt"),
                     func.sum(DealIntakeModel.arv).label("total_arv"))
            .group_by(DealIntakeModel.status).all()
        )
        stages = {
            r.status: {"count": r.cnt, "total_value": float(r.total_arv or 0)}
            for r in rows
        }
    finally:
        db.close()

    return jsonify(calc_pipeline_metrics(stages))


# ---------------------------------------------------------------------------
# Phase 5 — Seed case studies  (POST /api/seed/case-studies)
# ---------------------------------------------------------------------------

@app.route("/api/seed/case-studies", methods=["POST"])
def api_seed_case_studies():
    """
    Pre-load the 5 STLLC training case study properties.
    Idempotent — already-existing deal_ids are skipped.
    """
    from realtor_agent.core.seed_data import seed_case_studies
    data = request.get_json(force=True, silent=True) or {}
    run_uw = bool(data.get("run_underwriter", True))
    result = seed_case_studies(run_underwriter=run_uw)
    return jsonify({"success": True, **result})


# ---------------------------------------------------------------------------
# Phase 5 — Document upload/management
# ---------------------------------------------------------------------------

UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls",
    ".jpg", ".jpeg", ".png", ".gif", ".txt", ".csv",
}


@app.route("/api/deals/<int:deal_id>/documents", methods=["GET"])
def api_deal_documents_list(deal_id: int):
    """List all documents attached to a deal."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel, Document
    db = SessionLocal()
    try:
        deal = db.query(DealIntakeModel).get(deal_id)
        if not deal:
            return _json_error("Deal not found", 404)
        docs = db.query(Document).filter(Document.deal_id_fk == deal_id).all()
        return jsonify([_doc_to_dict(d) for d in docs])
    finally:
        db.close()


@app.route("/api/deals/<int:deal_id>/documents", methods=["POST"])
def api_deal_documents_upload(deal_id: int):
    """Upload a file and attach it to a deal. Multipart form-data."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel, Document
    import werkzeug.utils

    db = SessionLocal()
    try:
        deal = db.query(DealIntakeModel).get(deal_id)
        if not deal:
            return _json_error("Deal not found", 404)

        if "file" not in request.files:
            return _json_error("No file in request")

        f = request.files["file"]
        if not f.filename:
            return _json_error("No filename")

        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            return _json_error(f"File type {ext} not allowed")

        safe_name  = werkzeug.utils.secure_filename(f.filename)
        dest_dir   = UPLOADS_DIR / str(deal_id)
        dest_dir.mkdir(exist_ok=True)
        dest_path  = dest_dir / safe_name

        f.save(str(dest_path))
        file_size  = dest_path.stat().st_size

        doc = Document(
            deal_id_fk    = deal_id,
            document_type = request.form.get("document_type", "other"),
            file_name     = safe_name,
            file_path     = str(dest_path),
            file_size     = file_size,
            mime_type     = f.mimetype,
            uploaded_by   = request.form.get("uploaded_by", "web"),
        )
        db.add(doc)
        db.commit()
        db.refresh(doc)
        _push_notification(db, "Document uploaded",
                           f"{safe_name} attached to deal {deal.deal_id}",
                           "info", "deal", f"/deals/{deal_id}")
        db.commit()
        return jsonify({"success": True, "document": _doc_to_dict(doc)}), 201
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


@app.route("/api/documents/<int:doc_id>", methods=["DELETE"])
def api_document_delete(doc_id: int):
    """Delete a document record (and the file on disk if present)."""
    from realtor_agent.core.models import SessionLocal, Document
    db = SessionLocal()
    try:
        doc = db.query(Document).get(doc_id)
        if not doc:
            return _json_error("Document not found", 404)
        # Remove file from disk
        try:
            p = Path(doc.file_path)
            if p.exists():
                p.unlink()
        except Exception:
            pass
        db.delete(doc)
        db.commit()
        return jsonify({"success": True})
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


def _doc_to_dict(d) -> dict:
    return {
        "id":            d.id,
        "deal_id":       d.deal_id_fk,
        "document_type": d.document_type,
        "file_name":     d.file_name,
        "file_size":     d.file_size,
        "mime_type":     d.mime_type,
        "attorney_reviewed": d.attorney_reviewed,
        "uploaded_by":   d.uploaded_by,
        "created_at":    d.created_at.isoformat() if d.created_at else None,
    }


# ---------------------------------------------------------------------------
# Phase 5 — Bulk CSV / Excel import  (POST /api/deals/import)
# ---------------------------------------------------------------------------

@app.route("/api/deals/import", methods=["POST"])
def api_deals_import():
    """
    Bulk import from uploaded CSV or Excel (.xlsx) file.
    Required columns: address, contract_price, arv
    Returns {"inserted": int, "skipped": int, "errors": list}.
    """
    import csv, io, uuid
    from realtor_agent.core.models import SessionLocal, DealIntakeModel

    if "file" not in request.files:
        return _json_error("No file in request")

    f   = request.files["file"]
    ext = Path(f.filename).suffix.lower() if f.filename else ""

    rows: list[dict] = []

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower().replace(" ", "_")
                       for c in next(ws.iter_rows(max_row=1))]
            for xl_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, xl_row)))
        except Exception as exc:
            return _json_error(f"Excel parse error: {exc}", 400)
    else:
        # Treat as CSV
        try:
            content = f.read().decode("utf-8-sig")
            reader  = csv.DictReader(io.StringIO(content))
            rows    = [{k.strip().lower().replace(" ", "_"): v for k, v in r.items()}
                       for r in reader]
        except Exception as exc:
            return _json_error(f"CSV parse error: {exc}", 400)

    ALLOWED_IMPORT_FIELDS = {
        "address", "city", "state", "county", "zip_code", "asset_type",
        "contract_price", "arv", "rehab_cost", "monthly_rent",
        "taxes", "insurance", "hoa", "hold_months",
        "rate", "down_pct", "term_years",
        "vacancy_pct", "pm_pct", "repairs_pct", "capex_pct",
        "sell_cost_pct", "buy_close_pct",
        "acres", "zoning", "buildable", "road_access", "flood_zone",
        "status", "priority", "source", "assigned_to", "notes",
    }
    FLOAT_FIELDS = {
        "contract_price", "arv", "rehab_cost", "monthly_rent",
        "taxes", "insurance", "hoa", "rate", "down_pct", "acres",
        "vacancy_pct", "pm_pct", "repairs_pct", "capex_pct",
        "sell_cost_pct", "buy_close_pct",
    }
    INT_FIELDS   = {"hold_months", "term_years"}
    BOOL_FIELDS  = {"buildable", "road_access"}

    db = SessionLocal()
    inserted = skipped = 0
    errors   = []

    try:
        for i, row in enumerate(rows):
            if not row.get("address") or not row.get("contract_price"):
                skipped += 1
                continue
            try:
                kwargs = {"deal_id": f"IMP-{uuid.uuid4().hex[:8].upper()}"}
                for field in ALLOWED_IMPORT_FIELDS:
                    val = row.get(field)
                    if val is None or str(val).strip() == "":
                        continue
                    if field in FLOAT_FIELDS:
                        val = float(str(val).replace(",", "").replace("$", "").strip())
                    elif field in INT_FIELDS:
                        val = int(float(str(val).strip()))
                    elif field in BOOL_FIELDS:
                        val = str(val).lower() in ("1", "true", "yes")
                    kwargs[field] = val

                deal = DealIntakeModel(**kwargs)
                db.add(deal)
                db.commit()
                inserted += 1
            except Exception as exc:
                db.rollback()
                errors.append(f"Row {i+2}: {exc}")
    finally:
        db.close()

    return jsonify({"success": True, "inserted": inserted, "skipped": skipped, "errors": errors})


# ---------------------------------------------------------------------------
# Phase 5 — Extended deal PATCH  (all editable fields)
# ---------------------------------------------------------------------------
# (The existing /api/deals/<id> PATCH only allowed status/priority/notes/assigned_to.
#  Replace it with a fully-featured version.)

_DEAL_PATCH_ALLOWED = {
    # Pipeline/CRM
    "status", "priority", "notes", "assigned_to", "source",
    # Identity
    "address", "city", "state", "county", "zip_code", "asset_type",
    # Acquisition
    "contract_price", "arv", "rehab_cost", "dom",
    # Income
    "monthly_rent",
    # Operating expenses
    "taxes", "insurance", "hoa", "opex_other",
    # Expense ratios
    "vacancy_pct", "pm_pct", "repairs_pct", "capex_pct",
    # Exit/acquisition costs
    "sell_cost_pct", "buy_close_pct", "hold_months",
    # Financing
    "rate", "points", "down_pct", "term_years",
    # Land-specific
    "acres", "zoning", "buildable", "road_access", "flood_zone",
    # Profit targets
    "flip_profit_target_pct", "wholesale_min_fee", "investor_margin_pct",
    # Scoring
    "score", "deal_score",
}

_DEAL_FLOAT_FIELDS = {
    "contract_price", "arv", "rehab_cost", "monthly_rent",
    "taxes", "insurance", "hoa", "opex_other",
    "vacancy_pct", "pm_pct", "repairs_pct", "capex_pct",
    "sell_cost_pct", "buy_close_pct", "rate", "points", "down_pct",
    "acres", "flip_profit_target_pct", "wholesale_min_fee", "investor_margin_pct",
    "score", "deal_score",
}
_DEAL_INT_FIELDS = {"dom", "hold_months", "term_years"}
_DEAL_BOOL_FIELDS = {"buildable", "road_access"}


@app.route("/api/deals/<int:deal_id>/edit", methods=["PATCH"])
def api_update_deal_full(deal_id: int):
    """Full field PATCH — all DealIntakeModel editable fields."""
    from realtor_agent.core.models import SessionLocal, DealIntakeModel

    data = request.get_json(force=True) or {}
    updates = {}
    for k, v in data.items():
        if k not in _DEAL_PATCH_ALLOWED or v is None:
            continue
        if k in _DEAL_FLOAT_FIELDS:
            try:
                v = float(v)
            except (ValueError, TypeError):
                continue
        elif k in _DEAL_INT_FIELDS:
            try:
                v = int(float(v))
            except (ValueError, TypeError):
                continue
        elif k in _DEAL_BOOL_FIELDS:
            v = bool(v)
        updates[k] = v

    if not updates:
        return _json_error("No valid fields to update")

    db = SessionLocal()
    try:
        r = db.query(DealIntakeModel).filter(DealIntakeModel.id == deal_id).first()
        if not r:
            return _json_error("Deal not found", 404)
        for k, v in updates.items():
            setattr(r, k, v)
        r.updated_at = datetime.utcnow()
        db.commit()
        return jsonify({"success": True, "deal_id": deal_id, "updated": list(updates.keys())})
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Phase 5 — Deal-scoped activity log
# ---------------------------------------------------------------------------

@app.route("/api/deals/<int:deal_id>/activity", methods=["GET"])
def api_deal_activity_list(deal_id: int):
    """List activity events for a deal, most-recent first."""
    from realtor_agent.core.database import Activity
    from realtor_agent.core.models import SessionLocal, DealIntakeModel
    db = SessionLocal()
    try:
        deal = db.query(DealIntakeModel).get(deal_id)
        if not deal:
            return _json_error("Deal not found", 404)
        limit  = min(int(request.args.get("limit", 50)), 200)
        events = (
            db.query(Activity)
            .filter(Activity.deal_id_fk == deal_id)
            .order_by(Activity.created_at.desc())
            .limit(limit)
            .all()
        )
        return jsonify([_activity_to_dict(e) for e in events])
    finally:
        db.close()


@app.route("/api/deals/<int:deal_id>/activity", methods=["POST"])
def api_deal_activity_create(deal_id: int):
    """Log a new activity event on a deal."""
    from realtor_agent.core.database import Activity
    from realtor_agent.core.models import SessionLocal, DealIntakeModel
    data = request.get_json(force=True, silent=True) or {}
    if not data.get("activity_type"):
        return _json_error("activity_type is required")

    db = SessionLocal()
    try:
        deal = db.query(DealIntakeModel).get(deal_id)
        if not deal:
            return _json_error("Deal not found", 404)
        ev = Activity(
            deal_id_fk    = deal_id,
            activity_type = data["activity_type"],
            description   = data.get("description"),
            outcome       = data.get("outcome"),
            metadata_     = data.get("metadata"),
        )
        db.add(ev)
        db.commit()
        db.refresh(ev)
        return jsonify({"success": True, "activity": _activity_to_dict(ev)}), 201
    except Exception as exc:
        db.rollback()
        return _json_error(str(exc), 500)
    finally:
        db.close()


def _activity_to_dict(ev) -> dict:
    return {
        "id":            ev.id,
        "deal_id":       ev.deal_id_fk,
        "activity_type": ev.activity_type,
        "description":   ev.description,
        "outcome":       ev.outcome,
        "metadata":      ev.metadata_,
        "created_at":    ev.created_at.isoformat() if ev.created_at else None,
    }


# ---------------------------------------------------------------------------
# Training page route
# ---------------------------------------------------------------------------

@app.route("/training")
@app.route("/training/<lesson>")
def training_page(lesson: str | None = None):
    return render_template("training.html", lesson=lesson)


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return _json_error("Not found", 404)
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(e):
    logger.error("500 error: %s", e)
    if request.path.startswith("/api/"):
        return _json_error("Internal server error", 500)
    return "<h1>500 — Internal Server Error</h1>", 500


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    logger.info("Starting Realtor Agent on http://localhost:%d", port)
    app.run(host="0.0.0.0", port=port, debug=debug)
